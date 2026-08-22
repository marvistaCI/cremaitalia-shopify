# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\marvi\OneDrive\Pre-Vault\CremaItalia LLC\Operations\In USA\shopify\Systems\Systems Inventory.xlsx"

ARIAL = "Arial"
ESP = "FF55331B"
GOLD = "FFB88348"
HAIR = "FFD9D2C2"

hdr_fill = PatternFill("solid", fgColor=ESP)
hdr_font = Font(name=ARIAL, size=10, bold=True, color="FFFFFFFF")
cell_font = Font(name=ARIAL, size=10)
bold = Font(name=ARIAL, size=10, bold=True)
title_font = Font(name=ARIAL, size=14, bold=True, color=ESP)
sub_font = Font(name=ARIAL, size=9, italic=True, color="FF8C7E6A")
thin = Side(style="thin", color=HAIR)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
top_wrap = Alignment(vertical="top", wrap_text=True)
ctr = Alignment(vertical="top", horizontal="center", wrap_text=True)

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "Systems Inventory"

HEADERS = [
    "Category", "Name of System", "System URL", "Subscription Plan",
    "Monthly Cost Now (USD)", "Monthly Cost at Launch (USD)", "Other Costs",
    "Role Description", "IS_Shopify_App[Y/N]", "Currently_Subscribed[Y/N]",
    "Status", "Launch Blocking[Y/N]", "Owner / Lane", "Decision State",
    "Decision Source", "Notes",
]

R = []

R.append(["Commerce platform", "Shopify", "admin.shopify.com  /  crema-italia.myshopify.com",
 "Basic today; Grow recommended at launch", 39, 79,
 "Card rate is the real cost line: 2.9% + 30c (Basic), 2.7% + 30c (Grow), 2.5% + 30c (Advanced). Grow is $79/mo billed annually, $105 monthly.",
 "Storefront, catalog, cart, checkout, orders, customers, admin, hosted customer accounts.",
 "Y (platform)", "Y", "LIVE", "Y", "Steve / Code",
 "Platform LOCKED. Plan RECOMMENDED (Grow), Steve to confirm.",
 "production_build_spec.md section 10 checklist; CLAUDE.md sec9 2026-08-21",
 "Steve refers to the current plan as \"Pro\"; Shopify's 2026 ladder is Basic / Grow / Advanced / Plus (Grow was formerly named simply \"Shopify\"). Basic is disqualified rather than merely tight: it includes ZERO extra staff accounts and the team is Steve plus Lucia, Asia and Lauren. Grow allows 5. CORRECTION carried forward: Advanced break-even is ~$110k/month (~$1.3M/yr), not the $70-80k/yr an earlier note claimed. Plus (from $2,300/mo) was declined - about $24k/yr to hide one checkout field."])

R.append(["Payments", "Shopify Payments", "shopify.com/payments", "Included with the Shopify plan", 0, 0,
 "2.9% + 30c (Basic) / 2.7% + 30c (Grow) / 2.5% + 30c (Advanced). Chargebacks per Shopify schedule.",
 "Card processing and payouts to Mercury.",
 "Y (built-in)", "Y", "LIVE", "Y", "Steve",
 "Implied by the platform choice; fee schedule to confirm against the final plan.",
 "00_PROJECT_BRIEF.md sec18",
 "Not a subscription - the percentage IS the cost. On a SUBSCRIPTION order the all-in rate is Shopify 2.7% + Loop 1.0% = 3.7%, on the exact shelf the entire subscriber model rests on. This has never been checked against the pricing matrix (Store Operating Standards sec12.3)."])

R.append(["Subscriptions", "Loop Subscriptions", "loopsubscriptions.com  /  apps.shopify.com/loop-subscriptions",
 "Free today (50 subs); Starter at launch", 0, 99,
 "Starter adds 1.0% per transaction. Pro is $399/mo + 0.75% and is NOT needed. Fee is not monotonic - Pro's rate is lower; Starter to Pro breaks even at about $120k/month.",
 "Roccia subscription engine: selling plans, customer contracts, branded portal, dunning, cancellation flows, subscription-specific shipping rates.",
 "Y", "Y", "INSTALLED (dev store, Free tier)", "Y", "Steve / Code",
 "Engine LOCKED 2026-06-29. Tier RECOMMENDED (Starter), Steve to confirm.",
 "production_build_spec.md sec5.2.1, sec5.2.2; CLAUDE.md sec9 2026-08-21",
 "LOOP IS THE LARGER OF THE TWO PLATFORM COSTS and was locked for eight weeks before anyone priced it. Four named commitments in our own Standards sit above the Free tier: dunning management, cancellation flows, the branded portal, and subscription-specific shipping rates. VERIFIED IN THE APP, not off a pricing page: contract discount rates ARE editable on the FREE tier, so Pro's REST API is not required - promoting a Founding Member is a thirty-second manual admin edit, and the event is capped at 222 for the life of the business."])

R.append(["Reviews", "Judge.me", "judge.me  /  apps.shopify.com/judgeme", "Forever Free", 0, 0,
 "Awesome tier is $15/mo flat. NOT needed - Custom Questions was the only reason we would have bought it, and the palate join makes it unnecessary.",
 "Purchase-gated review collection, moderation queue and review-request emails. Writes Shopify's standard review metafields.",
 "Y", "Y", "INSTALLED (dev store, Free tier)", "N", "Steve / Code",
 "LOCKED as the collection/moderation backend 2026-08-20.",
 "production_build_spec.md sec6.1; Store Operating Standards sec13",
 "VERIFIED LIVE: product.metafields.reviews.rating and reviews.rating_count are readable by our own Liquid, server-side, no JavaScript - which is the whole input for our bespoke rating control and for aggregateRating. UNPROVEN, NOT REFUTED: reviews.product_reviews (individual records) returned nil on the dev store, probably because Judge.me syndicates review metaobjects through the Shop channel and a Partners dev store is not Shop-eligible. Open item C4. Configure: public storefront review form OFF (sec13.1), photo/video upload OFF (sec13.4). The vendor is deliberately swappable - it writes standard review data, so changing vendor changes the back office, not the storefront."])

R.append(["Banking", "Mercury", "mercury.com", "Free business checking", 0, 0,
 "Wire fees per Mercury's schedule; FX spread on EUR/SWIFT payments to Italian roasters.",
 "US business bank: roaster payments (EUR/SWIFT), vendor and people payments, Shopify payouts.",
 "N", "Y", "LIVE", "Y", "Steve",
 "CONFIRMED - account open.",
 "00_PROJECT_BRIEF.md sec2; Operations\\In USA\\Crema_Italia_Banking_Operations_Guide.odt",
 "Chosen specifically for Shopify integration plus EUR/SWIFT wire capability. Mercury does publish an API (read access and payment initiation) - it has never been evaluated or used here, and no integration is designed. The originating account for wires/SEPA is still listed as to-confirm in the brief's decisions-pending."])

R.append(["AI / build", "Anthropic Claude - Max", "claude.ai", "Max (5x $100/mo or 20x $200/mo)", 100, 100,
 "No annual discount and no trial on either Max tier. Anthropic API billed separately if ever used - it is not, today.",
 "Claude Code builds and deploys the theme and renders the Standards; Claude Cowork runs the OneDrive ops/brand lane.",
 "N", "Y", "LIVE", "Y (de facto)", "Steve",
 "In use. Which Max tier is active is not recorded.",
 "CLAUDE.md coordination callouts; docs/standards/collaboration-standard.md",
 "The entire two-lane working model depends on this subscription: Code owns the repo and all Standard sources, Cowork owns OneDrive and proposes rather than applies. $100 entered as the placeholder - confirm 5x vs 20x. Note this is currently a LARGER monthly spend than the Shopify plan."])

R.append(["Source control", "GitHub", "github.com/marvistaCI/cremaitalia-shopify", "Free (private repository)", 0, 0,
 "None. No paid features in use.",
 "Off-machine backup and full history for the Shopify theme repo and the three Standards.",
 "N", "Y", "LIVE", "N", "Code",
 "In use since 2026-06-23.",
 "CLAUDE.md sec9 2026-06-23",
 "Every POC batch is a commit, which is what makes preview-theme pruning safe - a deleted theme is redeployable from git. Private repo; no CI, no Actions, no paid seats."])

R.append(["Document store", "Microsoft OneDrive / Microsoft 365", "onedrive.com", "Personal or M365 (tier NOT recorded)", None, None,
 "Unknown - no document in the repo or in OneDrive records the tier or the cost.",
 "Operations, brand, legal and coordination documents; Cowork's working lane; read-only renders of the three Standards.",
 "N", "Y", "LIVE", "N", "Steve / Cowork",
 "In use. Cost and tier UNRECORDED.",
 "docs/standards/collaboration-standard.md sec9",
 "Path: C:\\Users\\marvi\\OneDrive\\Pre-Vault\\CremaItalia LLC. Holds the FDA pack, the roaster guides, the landed-cost model, the promissory-note register, the decisions log and the render-trust badge. A real business dependency with no recorded cost line - worth capturing. NOTE THE OVERLAP, now that Google Workspace is confirmed for email: Workspace includes Drive, so the business is potentially paying two vendors for cloud document storage. That is not automatically waste - the whole Cowork lane and every Standard render path is built on local OneDrive paths, and moving it would be real work for no functional gain. But it is worth a deliberate look rather than being discovered on a renewal notice."])

R.append(["Dev / test", "Shopify Partners development store", "partners.shopify.com  /  crema-italia-development.myshopify.com", "Free development store", 0, 0,
 "None.",
 "Platform-validation lab. Deliberately simulates the LOWEST plan we would launch on.",
 "Y", "Y", "LIVE", "N", "Code",
 "Created 2026-07-25 for the platform spike.",
 "CLAUDE.md sec9 2026-07-25",
 "Simulating Basic is deliberate: on Plus the checkout is customizable and the promo-field test would have returned a false pass. Both Loop and Judge.me install free here. CAVEAT carried forward: a dev store is not a perfect mirror of production - Shop-channel features are unavailable, which is the likely reason Judge.me's review metaobjects did not appear."])

R.append(["Fonts", "Google Fonts", "fonts.google.com", "Free", 0, 0, "None.",
 "Serves Marcellus (display) and Inter (body, including the real italic axis) to the storefront.",
 "N", "Y", "LIVE", "Y", "Code",
 "Locked with the 2026-07-01 artist rebrand.",
 "CLAUDE.md sec3.3; Brand Standards v2.2",
 "Theme loads from fonts.googleapis.com and fonts.gstatic.com; theme check raises RemoteAsset warnings for these, accepted. Montecatini Pro (the wordmark face) is a commercial Tipofili / Louise Fili font, outlined in the logo art, and deliberately never loaded as a webfont - which is why no font licence is being paid for."])

R.append(["Social", "Instagram - @cremaitaliaco", "instagram.com/cremaitaliaco", "Free business account", 0, 0,
 "Paid social is not budgeted anywhere.",
 "The only social profile linked from the storefront.",
 "N", "Y (link live in footer)", "LIVE", "N", "Steve",
 "Referenced by the theme; no content plan recorded.",
 "snippets/ci-footer.liquid:14",
 "The single outbound social link on the site. JSON-LD Organization.sameAs is deliberately NOT emitted yet, precisely because confirmed social profiles do not exist (production_build_spec.md sec9.3) - so this handle is asserted in the footer but not in structured data."])

R.append(["Domain / DNS", "Namecheap - cremaitalia.com (registrar + DNS)", "namecheap.com  /  ap.www.namecheap.com", "Domain registration (.com) + PremiumDNS add-on. Paid through 29-APR-2027.", None, None,
 "Namecheap .com renewal typically $15-20/yr; PremiumDNS roughly $5/yr. Both paid through 29-APR-2027.",
 "Registrar of record AND DNS host for cremaitalia.com. The zone points the domain at Shopify and carries the MX records that route company email to Google Workspace.",
 "N", "Y", "LIVE", "Y", "Steve",
 "REGISTRAR, DNS HOST AND RENEWAL DATE ALL CONFIRMED by Steve 2026-08-22.",
 "Steve, 2026-08-22 (this session). No prior document in the repo or OneDrive named any of it.",
 "RESOLVED 2026-08-22: nameservers FLIPPED to PremiumDNS (pdns1/pdns2.registrar-servers.com) and VERIFIED by querying the new nameservers directly rather than trusting the admin screen. Every record carried over intact and both nameservers agree: A 23.227.38.65 (Shopify), www CNAME shops.myshopify.com, MX smtp.google.com pref 1 (Google Workspace), the google-site-verification TXT, the full two-chunk 2048-bit DKIM key, and the DMARC record. Public resolvers (8.8.8.8, 1.1.1.1) already served the new delegation within minutes, and cremaitalia.com returned HTTP 200. No downtime, nothing lost. HISTORICAL CONTEXT - the finding that prompted the flip: PremiumDNS was subscribed through 29-APR-2027 while the domain's nameservers were set to Namecheap BASICDNS, so the premium service is not answering a single query. PremiumDNS only takes effect when the nameservers are pointed at pdns1/pdns2.registrar-servers.com. So there are exactly two honest options: switch the nameservers and start getting what is already paid for, or stop paying for it. What PremiumDNS actually buys over BasicDNS: a 100% uptime SLA (BasicDNS carries no guarantee), DDoS protection, and 30+ anycast nodes. DNSSEC IS NOT A DIFFERENTIATOR - Namecheap supports it on both, which is worth stating because it is the feature most people assume is the reason. This zone is unusually load-bearing: it simultaneously carries the storefront and the MX records for the address the FDA holds as the US Agent contact, so a DNS outage takes revenue and regulatory correspondence down in the same instant. Renewal 29-APR-2027 for both the domain and the add-on; confirm auto-renew is ON."])

R.append(["Email (mailboxes)", "Google Workspace - cremaitalia.com email", "workspace.google.com  /  admin.google.com", "Business Starter assumed - tier and seat count to confirm", None, None,
 "$7 per user/month on annual billing, $8.40 on flexible monthly (Business Starter). So 1 seat is $7/mo, 4 seats $28/mo. Seat count is the only missing input. EMAIL ALIASES ARE FREE - 30 per user, on every plan.",
 "Company email for cremaitalia.com. usagent@ is live and filed with the FDA. info@ and support@ do NOT exist yet.",
 "N", "Y", "LIVE - one alias missing", "Y", "Steve",
 "PROVIDER AND ALIASES CONFIRMED 2026-08-22 from the Workspace admin. Tier and seat count still to confirm.",
 "Steve, 2026-08-22 (this session); CLAUDE.md sec10; POC9 contact-form routing",
 "SETTLED 2026-08-22 FROM THE WORKSPACE ADMIN, and the project record was wrong. The primary user is steve.roberts@cremaitalia.com (created 22-JUN-2026), carrying FIVE aliases that all deliver into that one inbox: info@, sroberts@, steve@, usagent@ and roasters@. So info@ EXISTS and receives - CLAUDE.md sec10 says it does not, and that line should be corrected. THE EVIDENCE FOR THAT IS THE ADMIN SCREEN, NOT A REPLY - worth stating because a wrong inference was drawn first. A Shopify notification was sent TO steve@asymplat.biz FROM info@cremaitalia.com, and Steve replied to info@; that was briefly read as proof info@ receives, when it only proves mail was SENT to it. The alias existing in Workspace is what makes it receive, by construction. A conclusive empirical test is still one minute: mail info@ and support@ from an outside address and watch them arrive. usagent@ is confirmed live, which matters because it is the address on file with the FDA as the US Agent contact. WHAT IS ACTUALLY MISSING IS TWO ADDRESSES, NOT ONE: POC9's contact form routes More info to info@ (exists), I need help to support@ (DOES NOT EXIST) and Other to contact@ (DOES NOT EXIST). Both are free - 30 aliases per user on every plan. BEFORE CREATING THEM, TWO CALLS. (1) ALIAS OR GROUP? Every address today is an alias into Steve's single inbox. A Google Group is also free, needs no licence, and CAN INCLUDE EXTERNAL ADDRESSES - so support@ as a Group could reach Lauren without buying her a seat. Standard sec9 puts customer service with Crema Italia rather than the roaster, and the team is four people, so support@ is the one that most wants to be a Group. (2) DO THREE ADDRESSES EARN THEIR KEEP? info@, support@ and contact@ all landing in one person's inbox is three labels for one destination. The routing is worth keeping only if it will later fan out to different people; otherwise collapse the POC's three-way choice to two. HISTORICAL - what prompted the check. PARTIAL CONTRADICTION FOUND 2026-08-22 - live output beats the document, again. CLAUDE.md sec10 records that info@ and support@ do not exist, and this workbook repeated it. But a Shopify test notification sent that day arrived FROM info@cremaitalia.com, so info@ is at minimum configured as the Shopify SENDER address. THAT IS NOT THE SAME AS THE MAILBOX EXISTING. Sending and receiving are separate capabilities: Shopify can be told to send as any address at an authenticated domain, and the domain IS authenticated, so a working FROM line proves nothing about whether anything is listening on the other end. The MX records point at Google Workspace, so mail addressed TO info@ lands there and either finds a user, alias or group - or bounces. ONE-MINUTE TEST, STILL TO RUN: send a message from any outside address TO info@cremaitalia.com and see whether it arrives or bounces. Repeat for support@. Until that is done, treat receiving as unproven in BOTH directions rather than assuming either answer. THE MISSING MAILBOXES ARE FREE, AND THAT CHANGES THE ITEM. POC9's contact form routes More info to info@, I need help to support@ and Other to contact@; two of those three have never been created, which is why the form cannot ship and why JSON-LD contactPoint stays unemitted. On Google Workspace those addresses cost NOTHING - up to 30 aliases per user are included on every plan, so they are minutes in the admin console rather than new licences. ONE DESIGN CALL BEFORE CREATING THEM: an ALIAS delivers into one person's inbox, while a GOOGLE GROUP (also free, also no licence) is a shared inbox several people can work. Standard sec9 says customer service is handled by Crema Italia and never handed to a roaster, and the team is four people, so support@ probably wants to be a Group and info@ an alias. Decide that once, now, rather than migrating a busy address later."])

R.append(["Certificates", "Namecheap SSL certificate", "namecheap.com", "SSL certificate (tier to confirm)", None, None,
 "Namecheap SSL runs roughly $6-60/yr depending on tier. Cost and expiry to confirm from the dashboard.",
 "Purchased TLS certificate. NOT the certificate the storefront actually serves.",
 "N", "Y", "PAID - NOT IN USE on the storefront", "N", "Steve",
 "FOUND 2026-08-22. Steve mentioned the subscription; the live check contradicts its usefulness here.",
 "Live TLS check of cremaitalia.com, 2026-08-22; Shopify Help Centre",
 "ANSWERED 2026-08-22 BY READING THE NAMECHEAP ACCOUNT: IT IS ATTACHED TO NOTHING. PositiveSSL id 34236492, issued for cremaitalia.com, status INSTALLED, VALID TILL 12-NOV-2026, managed by the Namecheap SSL shared-hosting tool - and the account has NO hosting service at all (Hosting List reads: you do not have any Hosting service yet). So the certificate was validated for the domain and had nowhere to go; INSTALLED here means issued and activated, not deployed on a server. It cannot be on Shopify either, which refuses third-party certificates on every plan, and cremaitalia.com demonstrably serves a Shopify-issued Let's Encrypt certificate. A SECOND certificate, id 34236491, sits CANCELED and Not issued - two PositiveSSL orders, one cancelled before issuance and one issued with nowhere to live; worth a look at the order history for a possible double purchase. RECOMMENDATION: LET IT LAPSE 12-NOV-2026. The only thing that cannot be ruled out from inside Namecheap is whether the certificate was ever installed by hand on some non-Namecheap server - nothing suggests it was. DATE CORRECTION: an earlier note here said the whole Namecheap relationship is paid through 29-APR-2027. That covers the DOMAIN and PREMIUMDNS only; the certificate runs on its own clock and expires 12-NOV-2026. SECOND INSTANCE OF THE SAME PATTERN AS PREMIUMDNS - paid for, not serving. But this one CANNOT be switched on, which makes it a different decision. The certificate cremaitalia.com actually presents is issued by LET'S ENCRYPT (CN=cremaitalia.com, valid 07-JUL-2026 to 05-OCT-2026), auto-provisioned and auto-renewed by Shopify at its load balancer. SHOPIFY DOES NOT SUPPORT UPLOADING A THIRD-PARTY CERTIFICATE AT ALL - TLS is managed at the platform level for every store on every plan, which is why it is free. So a purchased certificate can never be installed on the Shopify storefront, now or after launch. TWO HONEST QUESTIONS: is this certificate doing a job somewhere else (another domain, a server, an internal tool)? If not, it is a renewable line item buying something Shopify already provides at no cost, and the money is better left unspent. Do NOT cancel before answering the first question - confirm what it is attached to."])

R.append(["Email deliverability", "SPF / DKIM / DMARC on cremaitalia.com", "admin.google.com  /  Namecheap Advanced DNS", "Part of Google Workspace + the DNS zone", 0, 0,
 "No licence cost. DNS records only - minutes of work, zero spend.",
 "The three records that decide whether mail from cremaitalia.com reaches inboxes rather than spam folders.",
 "N", "Y", "LIVE - SPF added and verified 2026-08-22", "N", "Steve",
 "GAP FOUND AND CLOSED 2026-08-22, same day.",
 "Live DNS queries against pdns1/pdns2 and public resolvers, 2026-08-22",
 "RESOLVED 2026-08-22: SPF ADDED AND VERIFIED. v=spf1 include:_spf.google.com ~all now resolves at the authoritative nameservers and at public resolvers, as exactly ONE spf1 record (the other apex TXT is the google-site-verification string, which is fine). Lookup budget is comfortable: _spf.google.com currently returns a FLAT record of ip4/ip6 ranges with no nested includes, so the whole apex SPF costs ONE of the permitted ten. VERIFIED ON A REAL MESSAGE 2026-08-22 - and the result exposes a distinction worth keeping. THERE ARE TWO INDEPENDENT SENDING PATHS AND THEY AUTHENTICATE DIFFERENTLY. (a) GOOGLE WORKSPACE mail - anything sent from info@, steve@, usagent@ and the rest - is what the APEX record governs. That is the path that was failing: Google's DMARC report showed source_ip 209.85.220.41, which falls inside 209.85.128.0/17 and is now covered. (b) SHOPIFY notifications authenticate via the MAILER SUBDOMAIN, not the apex - a live Shopify message shows spf=pass with smtp.mailfrom=mailera7q.cremaitalia.com and a Return-Path of bounces+...@mailera7q.cremaitalia.com. That path was ALWAYS passing and the apex record neither helped nor hindered it. So a Shopify message passing SPF does NOT validate the apex record, and vice versa - check the smtp.mailfrom domain before concluding which record a result is testing. Both paths now authenticate, by two different mechanisms. CORRECTION TO EARLIER ADVICE IN THIS ROW: adding Shopify Email will NOT consume apex SPF lookups. Shopify authenticates via a mailer SUBDOMAIN that carries its own SPF (see the Shopify Email row), so the apex record only ever needs Google. The 10-lookup caution was real but did not apply here. HISTORICAL - the evidence that prompted the fix. CONFIRMED BY GOOGLE'S OWN DMARC REPORT (aggregate report for 2026-08-19, supplied by Steve): DKIM pass, selector google, aligned to cremaitalia.com - and SPF result literally 'none', which DMARC then evaluates as spf FAIL. Three messages in that window. Disposition was 'none' only because the policy is p=none. THE GAP: THERE IS NO SPF RECORD AT ALL - zero v=spf1 records at the apex; the only apex TXT is a google-site-verification string. So DKIM alone is currently carrying every message, and it is a single point of failure - rotate a key, add a sender, or hit a path that breaks the signature and there is nothing underneath it. Fix is one TXT record at the apex: v=spf1 include:_spf.google.com ~all. Exactly ONE SPF record is permitted; two produces a permanent error and is worse than none. CORRECTION TO AN EARLIER ENTRY IN THIS WORKBOOK: a previous version claimed the DMARC reports 'go nowhere' because the RFC 7489 external-report authorisation record does not exist. The record genuinely does not exist - verified as NXDOMAIN under the correct name cremaitalia.com._report._dmarc.asymplat.biz, with a control query to prove the method - but Steve DOES receive the reports, so the conclusion drawn from it was wrong. Google sends them anyway. Adding the authorisation record is hardening, so that other reporters that do enforce the check also send, NOT the repair of a live failure. NOT CAUSED BY THE NAMESERVER FLIP: the two-chunk DKIM key survived it perfectly, as did DMARC and the verification TXT. SEPARATE LAUNCH TASK: Shopify Email sending as @cremaitalia.com needs its own authentication records in this same zone - do it in one pass with the SPF fix, and watch the SPF 10-lookup limit when adding a second include."])

R.append(["Telephony", "Dialpad", "dialpad.com", "Connect - Standard or Pro (tier and seat count to confirm)", None, None,
 "Connect Standard is $15 per user/month billed annually, $27 monthly. Pro is $25 annual, $35 monthly. Number porting and international calling are extra.",
 "Published business phone numbers. Carries the number filed with the FDA as the US Agent contact.",
 "N", "Y", "LIVE", "Y", "Steve",
 "CONFIRMED by Steve 2026-08-22. Tier and seat count to confirm.",
 "Steve, 2026-08-22 (this session). No prior document named the provider.",
 "COMPLIANCE-RELEVANT, WHICH IS EASY TO MISS FOR A PHONE SYSTEM. 00_PROJECT_BRIEF.md sec4 puts +1-813-376-4821 on file with the FDA as the US Agent contact for every partner roaster's Foreign Food Facility registration. If that is a Dialpad number, then this subscription lapsing does not merely lose calls - it breaks a regulatory contact of record, and the fix runs through the FDA rather than through a billing page. Worth confirming whether the FDA number is the Dialpad one or a personal mobile. NOTE THE OTHER DIRECTION TOO: the storefront currently publishes NO phone number anywhere - a grep of the theme finds no tel: link and no number in the footer, the contact page or the roaster records. POC9's contact form collects an optional phone FROM the customer but offers none back. So Dialpad is being paid for a capability the site does not yet use; either the site should publish a number or the seat count should reflect actual use."])

R.append(["Dev tooling", "Shopify CLI + Node.js", "shopify.dev/docs/api/shopify-cli", "Free", 0, 0, "None.",
 "theme pull / push / dev / check / list; the entire deploy and verification ritual.",
 "Y (tooling)", "Y", "LIVE", "N", "Code",
 "In use; authenticated on Steve's machine.",
 "CLAUDE.md sec7; .claude/skills/crema-poc-deploy",
 "The crema-poc-deploy skill wraps this: verify live state FIRST, validate, push, then prove the push by pull-and-diff. That ritual exists because on 2026-07-24 a stale document was trusted over a live check and produced a duplicate theme."])

R.append(["Automation", "Shopify Flow", "apps.shopify.com/flow", "Free (Shopify first-party)", 0, 0, "None.",
 "Customer tagging (founding-member-NNN, subscription state), the Offerta aging transition, pallet-gap triggers.",
 "Y", "N", "SELECTED - not yet installed", "Y", "Code",
 "LOCKED as the automation layer.",
 "Store Operating Standards sec11, sec12.2",
 "NEVER TESTED - open item C5 on the round-2 validation list. Liquid date comparison is reported unreliable inside Flow, so assume a Run-code step. The entitlement tags the discount engine reads are maintained HERE plus Loop webhooks, never by the theme, because a customer can cancel from an email link and never touch storefront UI."])

R.append(["Bundles", "Shopify Bundles (native)", "apps.shopify.com/shopify-bundles", "Free (Shopify first-party)", 0, 0, "None.",
 "Sorpresa collections as composite BOM products, with component stock kept in sync.",
 "Y", "N", "SELECTED - not yet installed", "Y", "Code",
 "RECOMMENDED 2026-08-21 over any paid bundle app.",
 "production_build_spec.md sec7.1; Store Operating Standards sec7",
 "Third-party bundle apps earn their fee on mix-and-match, build-your-own, volume discounts and BOGO - none of which we need, because a Sorpresa collection is a FIXED set of components we choose. Native limits (100 variants, 30 products, 3 option dimensions) are nowhere near binding. The two requirements NO app satisfies are ours to build either way: component-derived facets, and availability gated on component FRESHNESS. OPEN (item B2, ten minutes): does native Bundles actually decrement COMPONENT inventory? Sources conflict; the recommendation changes if it does not."])

R.append(["Email marketing", "Shopify Email", "apps.shopify.com/shopify-email", "Free tier at launch", 0, 0,
 "Free monthly email allowance, then per-email pricing above it.",
 "The launch email flows: welcome, abandoned cart, welcome series, Roccia lifecycle, pallet-gap, Selezione early access, win-back.",
 "Y", "Partial - domain authentication ALREADY in the DNS zone", "SELECTED for launch", "Y", "Steve",
 "DECIDED for launch; migrate to Klaviyo at a trigger not yet confirmed.",
 "00_PROJECT_BRIEF.md sec10, sec15; DNS zone inspection 2026-08-22",
 "CORRECTION TO AN EARLIER ENTRY: Shopify sender authentication is NOT a future launch task - it is already in the zone, and VERIFIED HEALTHY END TO END on 2026-08-22. TWO complete sets exist and BOTH work. Set A: mailera7q -> 34e4bb34ae83.p662.email.myshopify.com (SendGrid), DKIM hosts a7q._domainkey and a7q2._domainkey, both chasing through to live RSA keys. Set B: mailerast -> 71e720ca2d24.p581.email.myshopify.com (Mailgun), DKIM hosts pdk1._domainkey.mailerast and pdk2._domainkey.mailerast - NESTED UNDER THE MAILER SUBDOMAIN, which is the detail that matters - also chasing through to live RSA keys. Nothing is broken and nothing needs fixing. SET A IS THE LIVE SIGNING IDENTITY, CONFIRMED 2026-08-22 from the headers of a real Customer account activation message: DKIM-Signature carries s=a7q and d=cremaitalia.com, the Return-Path is bounces+...@mailera7q.cremaitalia.com, and the relay is o31.mailer.shopify.com (SendGrid) - three independent signals agreeing. Authentication-Results on that message: dkim=pass (2048-bit), spf=pass, dmarc=pass, and a receiving spam score of 0 with delivery to the inbox. DO NOT DELETE SET B ON THAT EVIDENCE ALONE: one message proves a7q signed THAT notification, not that Shopify never uses the Mailgun path for a different message class, and transactional and marketing sends do not always share infrastructure. The cheap confirmation is Shopify admin, Settings > Notifications > Sender email, which should name the identity it considers active - remove the other only if it names just one. READ FROM THE SHOPIFY ADMIN 2026-08-22 (Settings > Notifications): Sender email is info@cremaitalia.com and Email domain authentication reads AUTHENTICATED, 'DNS records updated globally'. Because authentication is in place, mail goes out as info@cremaitalia.com rather than falling back to the store+73699983529@shopifyemail.com address Shopify reserves for stores whose authentication or DMARC is not set up. THE ADMIN CANNOT SETTLE THE SET A vs SET B QUESTION, AND THAT IS WORTH RECORDING SO NOBODY LOOKS AGAIN: Shopify HIDES the expected DNS record list once a domain is verified. The entire content of the email-domain page is the words 'Email domain authentication / Authenticated / DNS records updated globally' - verified by walking the full shadow DOM, since the admin is built from web components and plain text extraction returns almost nothing. There is no table naming a7q or pdk to compare against the zone. THE REMAINING QUESTION IS TIDINESS, NOT FAULT: two sending identities are authenticated against this domain, one per Shopify sending provider generation, and it is worth knowing which belongs to the production store and which to the development store before launch - purely so that a launch-day deliverability question does not start with an hour of working out what these records are. TWO WRONG CONCLUSIONS WERE DRAWN HERE BEFORE THE RIGHT ONE, both worth recording. (1) The DKIM hosts were GUESSED as pdk1._domainkey, which returned NXDOMAIN, and that absence was read as a fault; the real host was nested under mailerast and was simply never tried. (2) A check for whether the CNAME targets served keys grepped for the literal string DKIM1 - which matched the HOSTNAME dkim1 in Set A's output and failed against dkim3/dkim4, reporting Set B's healthy targets as dead. It matched text rather than meaning, and returned the exact opposite of the truth. Same failure class this project has logged before (document.fonts.check returning true for synthesised faces; a truthy empty Liquid drop; case-sensitive matching defeated by text-transform). The fix that worked was matching p=MII, the actual key material. NOTE THE USEFUL CONSEQUENCE: because Shopify puts its SPF on the mailer SUBDOMAIN rather than the apex, it costs nothing against the apex SPF lookup limit. HARD CONSTRAINT for whichever platform is used: campaign discounts are delivered as TIME-BOXED CUSTOMER TAGS, never as codes and never as /discount/ links (Store Operating Standards sec3 - a discount link carries a real, readable code, which is the leak Steve identified). The email links to the store or the cart; the rate is already attached to that customer. Flow 8 (referral milestones) stays UNBUILT until the referral reward is decided."])

R.append(["Access control", "Locksmith", "apps.shopify.com/locksmith  /  locksmith.guide", "Middle tier (about $12/mo suggested)", 0, 12,
 "\"Pay what feels good\" pricing - the shown price is a suggestion, roughly a third of your Shopify plan cost. A free plan and a 15-day trial exist.",
 "Tag-gates the Selezione shelf to active-roccia customers for the first 48 hours after a new SKU drops.",
 "Y", "N", "SELECTED - but VERIFY still needed", "N", "Steve / Code",
 "Decided in the 2026-06 brief. Never re-tested against the POC or the current Standards.",
 "00_PROJECT_BRIEF.md sec10, sec11; Store Operating Standards sec1",
 "This is the OLDEST app decision in the whole record and it predates the POC entirely. The requirement is still live - the 48-hour early-access rule survives in Store Operating Standards sec1 - but nobody has checked whether a customer tag plus native collection availability, or a Shopify Function, could do the same job without a paid app. Worth one look before it becomes a standing $12/mo."])

R.append(["Discounts", "Shopify Functions (custom app, in-house)", "shopify.dev/docs/apps/build/discounts", "Free on ALL plans; we build and host it", 0, 0,
 "Developer time only. Not plan-gated - Functions replaced the old Plus-only Scripts.",
 "A single evaluator returning MAX() of every discount a customer qualifies for. No stacking, ever.",
 "Y (custom)", "N", "BUILD - A1 answered, shape now known", "Y", "Code",
 "POLICY LOCKED (MAX, no stacking). Mechanism: A1 and sec12.7 answered 2026-08-22; sec11/sec12.8 ownership still Steve's call.",
 "Store Operating Standards sec11, sec12.7, sec12.8; production_build_spec.md sec5.2; open items A1/A2",
 "A1 ANSWERED EMPIRICALLY 2026-08-22 (commit fa07ad3), and it changes the shape of this in our favour. A throwaway Discount Function was built, deployed to the dev store and driven through a real cart and checkout. Same $24.95 variant twice: on Loop's 12% plan the Function's 10% came off the ALREADY-REDUCED $21.96 and billed $19.77, an effective 20.76%; the one-time control billed $22.46, a clean 10%. combinesWith was false on all three discount classes and made NO difference, because a selling-plan adjustment is not a discount and never enters the combination contest. SO IT COMPOUNDS - but the useful half is the second finding: the Function CAN see the subscription line and IS handed the pre-plan base price, so it can apply a TOP-UP to MAX rather than being barred from subscription lines. That is what stops MAX silently collapsing to the standing rate on every subscription. THE EITHER/OR FRAMING IS SUPERSEDED - the rate no longer has to live in the selling plan OR in a Function. Standard sec12.7 fell out of the same run and is ANSWERED YES: a discount Function can read customer tags and metafields. Recommendation only - Standard sec11/sec12.8 remains Steve's decision. HISTORICAL - what the spike had established before that test. THE ONE ARCHITECTURAL BREAK THE SPIKE FOUND. Discount Functions do NOT re-run when recurring orders are created - the rate is snapshotted onto the Loop subscription contract at signup, and orders 2..n bill from that snapshot. So a Function CANNOT own subscription entitlement as Standard sec11 specifies: someone who subscribes at 10% and later becomes a founder would keep 10% forever, and the 60-day benefit grace could not be enforced. The rate must live on the CONTRACT, which makes it Loop's job, and shrinks the Function to campaign discounts on one-time purchases. A1 is the last blocking question: does a Function COMPOUND with the selling-plan price adjustment on the first order? Until answered, the subscriber rate lives in the selling plan OR in a Function, never both."])

R.append(["Accounts", "Customer-account UI extension (custom, in-house)", "shopify.dev/docs/api/customer-account-ui-extensions", "Free on ALL plans; we build it", 0, 0,
 "Developer time only.",
 "The account page: membership tile, founder number, taste profile, Loop portal slot, order history.",
 "Y (custom)", "N", "BUILD - never prototyped", "Y", "Code",
 "De-risked 2026-08-21. Research, not yet a spike.",
 "production_build_spec.md sec5.1; open item C1",
 "DE-RISK: extensions run on ALL plans, not just Plus, so the account experience does not push the plan decision. They can READ AND WRITE customer metafields, which confirms taste-profile-as-customer-metafield is the natural mechanism rather than a workaround. THE COST IS BRAND, NOT FUNCTION: no custom CSS, no arbitrary HTML, no custom fonts - only Shopify's component library under the shared branding configuration. The POC's account information architecture and copy survive; its visual design does not. Two-minute check outstanding (B1): does the branding editor offer Marcellus?"])

R.append(["Shipping", "Shopify Shipping (USPS / UPS)", "shopify.com/shipping", "Included with the Shopify plan", 0, 0,
 "Per-label postage. Customer pays $8.50 flat under $55; free at $55+ and on every Roccia shipment.",
 "Label purchase and rates: USPS Ground Advantage under 1 lb, UPS Ground at 1 lb and above.",
 "Y", "N", "SELECTED", "Y", "Steve",
 "Carriers decided. UPS account setup status is OPEN.",
 "00_PROJECT_BRIEF.md sec13, sec18; Store Operating Standards sec8",
 "No FedEx and no expedited at launch. US only (50 states + DC + territories). The Shopify PACKING SLIP TEMPLATE must carry no monetary fields at all - Standard sec8.1 says nothing inside any package shows a price, gift or not, and the receipt is an email entitlement. Verify that template after any theme or settings change."])

R.append(["Search", "Shopify Search & Discovery", "apps.shopify.com/search-and-discovery", "Free (Shopify first-party)", 0, 0, "None.",
 "Native predictive search and filtering on the production storefront.",
 "Y", "N", "CANDIDATE", "N", "Code",
 "Noted as the production route when the POC's fake search was removed.",
 "CLAUDE.md sec9 2026-07-17 (POC9)",
 "The POC's header search icon was REMOVED in POC9 because it promised a search field and only routed to the Shop page. Production search was explicitly deferred to this native app. Also relevant to JSON-LD: WebSite.potentialAction / SearchAction is deliberately not emitted until real search ships."])

R.append(["Forms", "Shopify Forms", "apps.shopify.com/shopify-forms", "Free (Shopify first-party)", 0, 0, "None.",
 "Newsletter and email capture in the footer and on the coming-soon page.",
 "Y", "N", "CANDIDATE - undecided", "N", "Code",
 "OPEN - no decision recorded.",
 "POC14 (footer email capture); coming-soon page signup form",
 "POC14 added footer email capture to the theme because the storefront had none while the coming-soon page it replaces has one. Production needs a real destination for those addresses - Shopify Forms, Shopify Email, or Klaviyo. Nothing decides this yet, and it is cheap to get wrong (addresses captured into nothing)."])

R.append(["Email marketing (later)", "Klaviyo", "klaviyo.com", "Email plan, priced by ACTIVE PROFILES", 0, 0,
 "2026 pricing: about $20/mo at 500 profiles, $30 at 1,000, $150 at 10,000. SMS billed separately. Billed monthly, no annual discount.",
 "Successor to Shopify Email once the flows outgrow it.",
 "Y", "N", "CANDIDATE - migration trigger unconfirmed", "N", "Steve",
 "PENDING - trigger named in the brief's decisions-pending.",
 "00_PROJECT_BRIEF.md sec10, sec18",
 "The brief estimated $35-45/mo; 2026 pricing moved to ACTIVE-PROFILE billing (Feb 2025) and starts lower at our expected launch list size. Default migration trigger is about 200 active subscribers, or the first automation Shopify Email cannot express (win-back, anniversary and pallet-gap were the named ones). Loop Starter integrates with Klaviyo natively, which matters if subscription lifecycle emails move there."])

R.append(["Customer service", "Shopify Inbox  vs  Gorgias  vs  email only", "apps.shopify.com/inbox", "TBD", 0, 0,
 "Shopify Inbox is free. Gorgias is typically $10-60/mo at our volume.",
 "Customer service channel. English only, handled by Crema Italia - a US customer is never put in front of a partner roaster.",
 "Y (if Inbox)", "N", "PENDING DECISION", "N", "Steve",
 "OPEN - an explicit three-way choice in the brief.",
 "00_PROJECT_BRIEF.md sec18, sec9",
 "Whichever is chosen has to receive POC9's contact-form routing (More info / I need help / Other, to three different mailboxes) - and two of those mailboxes do not exist yet, so this decision is coupled to the mailbox gap above."])

R.append(["Affiliates", "Refersion  /  UpPromote  /  GoAffPro", "apps.shopify.com (candidates)", "TBD", 0, 0,
 "Typically $30-100/mo depending on vendor and tier.",
 "Commission tracking, referral links and payouts for the footer Affiliates program.",
 "Y", "N", "CANDIDATE - deferred post-launch", "N", "Steve",
 "DEFERRED to post-launch.",
 "production_build_spec.md sec4",
 "The theme owns ONLY the landing page and the application entry point; commission tracking, links and payouts are all app tooling. Deliberately distinct from About > Partners (people we actually work with, such as the freight forwarder) and from the sec12.6 referral program, which is a different mechanic aimed at customers."])

R.append(["Referral", "Referral capture and reward tooling", "TBD", "TBD", None, None,
 "Unknown - the reward FORM is itself undecided.",
 "Capture, attribution and reward for customer referrals.",
 "N", "N", "PENDING DECISION", "N", "Steve",
 "OPEN - Store Operating Standards sec12.6.",
 "Store Operating Standards sec3 (Referral = TBD), sec12.6",
 "The old \"free 100g bag\" reward is VOID: 100g exists only inside Sorpresa collections, so there is no standalone SKU to gift. Both the reward form (a 250g bag? account credit?) and the capture tooling are open. HARD CONSTRAINT added in v1.3: whatever is chosen must NOT require issuing a discount code, because sec3 bans codes outright."])

R.append(["Fulfilment", "Tampa 3PL", "NOT SELECTED", "n/a", None, None,
 "Receiving $/pallet, storage $/pallet/month, pick-pack $/order + $/extra unit, returns handling, minimums - none of it quoted yet.",
 "Receiving, food-grade storage, pick-pack, FIFO enforcement, printed inserts.",
 "N", "N", "NOT SELECTED", "Y", "Steve",
 "OPEN - Store Operating Standards sec12.9.",
 "Store Operating Standards sec8, sec8.1, sec12.9; 00_PROJECT_BRIEF.md sec9, sec18",
 "THREE QUALIFYING QUESTIONS, not preferences, all to be asked before commercial terms. (0) How would you segregate an Offerta lot from FRESH stock of the SAME physical SKU? A SKU carries no shelf segment, so an Offerta split makes a second Shopify product drawing on the same bin - and FIFO would hand the aged bag to a full-price buyer. This one is unsolved on our side too. (a) Do you print OUR packing slip or insert your own paperwork? Own paperwork breaks sec8.1 and nothing on our side prevents it. (b) Can you insert a printed card varied per order? A 3PL that cannot do this cannot fulfil Sorpresa at all, because those collections already ship a tasting card. Study: Operations\\Fulfillment Selection\\Tampa_3PL_Selection.docx."])

R.append(["Freight", "Italian freight forwarder", "NOT SELECTED", "n/a", None, None,
 "Per-shipment air freight, FCA Incoterm, roastery-to-airport haul included in our cost.",
 "Roastery to airport haul, then FCO or MXP to Tampa (TPA) air freight.",
 "N", "N", "NOT SELECTED", "Y", "Lucia / Steve",
 "OPEN.",
 "00_PROJECT_BRIEF.md sec8; Operations\\Fulfillment Selection\\Italian_Freight_Forwarder_Selection.docx",
 "Air freight only. Carriers seen: American Airlines FCO-TPA direct, Lufthansa Cargo via FRA, IAG Cargo via LHR, Delta via ATL/JFK, SWISS via ZRH. \"Partner 1 - Freight Forwarder - Italy\" is the placeholder card on the storefront About page, so this selection also unblocks a piece of site content."])

R.append(["Customs", "US customs broker", "NOT SELECTED", "n/a", None, None,
 "Per-entry fees. MPF 0.3464% applies; HMF is ocean-only so not applicable to air. DUTY IS 0% on roasted coffee.",
 "Entry filing, ISF where applicable, FDA Prior Notice, and FDA hold management.",
 "N", "N", "NOT SELECTED", "Y", "Steve",
 "OPEN.",
 "00_PROJECT_BRIEF.md sec8; Operations\\Fulfillment Selection\\US_Customs_Broker_Selection.docx",
 "HS codes: 0901.21 roasted not decaffeinated, 0901.22 roasted decaffeinated. Duty 0%. Documents travelling with the pallet: English commercial invoice, English packing list one line per SKU per bag size, ICO Certificate of Origin, ASL health certificate if applicable. No phytosanitary certificate is required for roasted coffee."])

R.append(["Compliance", "FDA Industry Systems (FURLS / Prior Notice)", "access.fda.gov", "Government system - no subscription", 0, 0,
 "No FDA fee for facility registration. Broker fees for Prior Notice filing may apply.",
 "Foreign Food Facility registration for every partner roaster; Prior Notice filed before every shipment.",
 "N", "Y (in use)", "LIVE", "Y", "Steve / Lucia",
 "Required by 21 CFR Part 1, Subpart H.",
 "00_PROJECT_BRIEF.md sec4; Operations\\In Italy\\Crema_Italia_FDA_Quick_Start_v5.pdf",
 "Crema Italia provides US Agent service free of charge; the agent on file with the FDA is Steve Roberts, usagent@cremaitalia.com. Biennial renewal per roaster. Prior Notice is filed by us, per shipment, from roaster-supplied data. Not a paid system, but it is a hard dependency and belongs in the ecosystem picture."])

R.append(["Accounting", "QuickBooks  vs  Wave  vs  spreadsheets", "quickbooks.intuit.com  /  waveapps.com", "TBD", 0, 0,
 "QuickBooks is typically $35-100/mo. Wave has a free tier. Spreadsheets are $0.",
 "Bookkeeping, roaster payables, landed-cost postings, promissory-note register.",
 "N", "N (spreadsheets in use)", "PENDING DECISION", "N", "Steve",
 "OPEN - both names appear once each in the brief, neither chosen.",
 "00_PROJECT_BRIEF.md; Operations\\Accounting\\",
 "Today accounting runs entirely on OneDrive spreadsheets: Accounting_Log.xlsx, Promissory_Notes_Register.xlsx and Crema_Italia_Landed_Cost_Model_v1.xlsx. Nothing is integrated with Mercury or with Shopify. That is fine at pre-launch volume and will not be once orders and roaster payables both start moving."])

R.append(["Pricing tool", "SKU price-maintenance engine (phased, in-house)", "n/a", "Phase 1 spreadsheet; Phase 3 custom app", 0, 0,
 "Phase 1 is free (spreadsheet). Phase 3 is developer time, unscoped.",
 "Landed cost x markup with approval governance. NOT a native Shopify feature.",
 "N", "Partial (the spreadsheet is in use)", "PHASED - approach LOCKED", "Y (phase 1)", "Steve / Code",
 "LOCKED 2026-07-13 as a phased approach.",
 "Store Operating Standards sec11, sec12.2, sec2.4",
 "Phase 1 at launch: compute in Crema_Italia_Landed_Cost_Model_v1.xlsx and enter prices by hand in Shopify; governance is admin review at entry time. Adequate for a small catalog on a 6-10 week lot cadence. Phase 2: Shopify Flow for the Offerta AGING transition (date-triggered shelf move, price recalc to the O[size] factor, admin alert). Phase 3: a lightweight approve/hold/defer queue app WHEN volume justifies it. Back office only - it does not block or shape the theme build."])

R.append(["Analytics", "Shopify Analytics / Google Analytics (GA4)", "analytics.google.com", "Free", 0, 0, "None at our volume.",
 "Traffic, funnel and conversion reporting.",
 "N", "N", "CANDIDATE - no decision recorded", "N", "Steve",
 "OPEN - nothing in the repo or the Standards decides an analytics stack.",
 "Mentioned only in passing; no governing document",
 "IMPORTANT CONSEQUENCE OF THE LOOP FINDING: Shopify's discount analytics will report ZERO discounts on subscription orders, because a selling-plan discount is a PRICE ADJUSTMENT rather than a discount - test order #1001 billed $24.95 as $21.96 with no discount line anywhere. So subscriber-benefit reporting cannot come from Shopify's discount reports, and the theme has to render the benefit itself or the customer never learns they got it (open item A3)."])

R.append(["Research", "Perplexity", "perplexity.ai", "Tier not recorded", None, None, "Unknown.",
 "Market, regional and competitor research.",
 "N", "Unknown", "UNRECORDED", "N", "Steve",
 "-",
 "OneDrive\\Crema_Italia_Memory_Document_from_Perplexity.md",
 "Appears once as the source of a research document, plus a Perplexity-derived Tuscan roasters purchasing-agent summary in Operations\\In Italy. Whether there is an active paid subscription is recorded nowhere. Listed for completeness rather than because it is a dependency."])

R.append(["Data jobs", "Reorder-rate and palate-match computation (in-house)", "n/a", "Build - no runtime chosen", 0, 0,
 "Hosting plus developer time. No host has been selected.",
 "Scheduled computation writing reorder rate and palate-matched feedback into product metafields.",
 "N", "N", "BUILD - not started", "N", "Code",
 "Decided 2026-08-20 as part of the trust and social-proof work.",
 "production_build_spec.md sec6.1; Store Operating Standards sec13.6",
 "Neither figure can be derived in Liquid. Both must honour a minimum-n floor with SILENCE below it, and the floor must be a named constant, never a literal in a template. Bottega is excluded from reorder rate BY SHELF, not by the floor - nobody rebuys a grinder, so the number would sit near zero, mean nothing and read as damning. NOTE: no runtime has been chosen for this job, which is a quiet infrastructure decision hiding inside a content feature. CANDIDATE ADDED 2026-08-22, and it is the cheapest one available: a SCHEDULED GITHUB ACTION. The repo already exists, is already paid for at $0, and Code already maintains it - so a scheduled workflow calling the Admin API adds no vendor, no host and no new credential store. That is a better fit than standing up a web host for something that is a periodic script rather than a service."])

ws["A1"] = "Crema Italia, LLC - Systems Inventory"
ws["A1"].font = title_font
ws["A2"] = ("Every system, app, platform and vendor referenced, decided or left pending across the theme repo, the three Standards, "
            "the project brief, the OneDrive operations tree and the Claude Code session record. Research only - no source document was changed to produce this.")
ws["A2"].font = sub_font
ws["A3"] = "Compiled 2026-08-22.  Costs are USD per month.  A blank cost cell means genuinely unknown or not applicable, never zero-by-assumption.  Read the Notes column before quoting any figure."
ws["A3"].font = sub_font

for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = box

for row_vals in R:
    ws.append(row_vals)

widths = [18, 30, 34, 30, 13, 14, 34, 40, 13, 15, 22, 12, 15, 30, 32, 80]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

last = 4 + len(R)
for row in ws.iter_rows(min_row=5, max_row=last, min_col=1, max_col=len(HEADERS)):
    for cell in row:
        cell.font = cell_font
        cell.alignment = top_wrap
        cell.border = box
    row[1].font = bold
    for j in (8, 9, 10, 11):
        row[j].alignment = ctr
    for j in (4, 5):
        row[j].number_format = '$#,##0;($#,##0);-'
        row[j].alignment = Alignment(vertical="top", horizontal="right")

STATUS_FILL = {
    "LIVE": "FFEFF6EF",
    "INSTALLED": "FFEFF6EF",
    "SELECTED": "FFFDF6EA",
    "BUILD": "FFFDF6EA",
    "PHASED": "FFFDF6EA",
    "CANDIDATE": "FFF6F3EC",
    "PENDING": "FFFBEDEC",
    "NOT SELECTED": "FFFBEDEC",
    "GAP": "FFFBEDEC",
    "UNRECORDED": "FFF6F3EC",
}
for r in range(5, last + 1):
    st = str(ws.cell(row=r, column=11).value or "")
    for k, v in STATUS_FILL.items():
        if st.startswith(k):
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=v)
            break
    ws.row_dimensions[r].height = 105

ws.freeze_panes = "C5"
ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(HEADERS)), last)

DATA_FIRST, DATA_LAST = 5, last

cs = wb.create_sheet("Cost Summary")
cs["A1"] = "Monthly cost roll-up"
cs["A1"].font = title_font
cs["A2"] = "Every figure pulls from the Systems Inventory sheet. Change a cost there and this recalculates."
cs["A2"].font = sub_font

cs["A4"] = "Line"; cs["B4"] = "Monthly (USD)"; cs["C4"] = "Annual (USD)"; cs["D4"] = "What it covers"
for c in range(1, 5):
    cc = cs.cell(row=4, column=c)
    cc.font = hdr_font; cc.fill = hdr_fill; cc.border = box
    cc.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

rng_now = "'Systems Inventory'!$E$%d:$E$%d" % (DATA_FIRST, DATA_LAST)
rng_launch = "'Systems Inventory'!$F$%d:$F$%d" % (DATA_FIRST, DATA_LAST)

lines = [
    ("Subscribed and being paid today", "=SUM(%s)" % rng_now,
     "Everything already running: Shopify Basic, Claude Max, Mercury, GitHub, Google Fonts, Loop Free, Judge.me Free and the dev store."),
    ("Platform floor at launch (Shopify Grow + Loop Starter)", "=79+99",
     "The two numbers that matter. Stated in production_build_spec.md sec5.2.1 as $178/mo before card fees."),
    ("All DECIDED systems at launch", "=SUM(%s)" % rng_launch,
     "Adds Locksmith and keeps Claude Max. Excludes every PENDING and NOT SELECTED line, which is where the real remaining uncertainty sits."),
    ("Not yet costed at all", None,
     "3PL, freight forwarder, customs broker, accounting, customer service, referral tooling, domain and DNS, mailboxes, and the OneDrive tier. Several of these are launch-blocking."),
]
r = 5
for label, formula, cover in lines:
    cs.cell(row=r, column=1, value=label).font = bold
    if formula:
        cs.cell(row=r, column=2, value=formula).number_format = '$#,##0'
        cs.cell(row=r, column=3, value="=B%d*12" % r).number_format = '$#,##0'
    else:
        cs.cell(row=r, column=2, value="n/a")
        cs.cell(row=r, column=3, value="n/a")
    cs.cell(row=r, column=4, value=cover)
    for c in range(1, 5):
        cell = cs.cell(row=r, column=c)
        if c != 1:
            cell.font = cell_font
        cell.alignment = top_wrap
        cell.border = box
    cs.row_dimensions[r].height = 46
    r += 1

r += 1
cs.cell(row=r, column=1, value="Variable costs - deliberately NOT included in the totals above").font = Font(name=ARIAL, size=11, bold=True, color=ESP)
r += 1
var = [
    ("Shopify Payments card rate", "2.9% + 30c on Basic; 2.7% + 30c on Grow; 2.5% + 30c on Advanced."),
    ("Loop transaction fee", "0% on Free, 1.0% on Starter, 0.75% on Pro."),
    ("All-in rate on a SUBSCRIPTION order at launch", "3.7% (Shopify 2.7% + Loop 1.0%). Never run against the pricing matrix - Store Operating Standards sec12.3."),
    ("Outbound postage", "USPS Ground Advantage under 1 lb; UPS Ground at 1 lb+. Customer pays $8.50 flat under $55, free at $55+ and on every Roccia shipment."),
    ("Inbound air freight, customs, MPF", "Per shipment. Duty is 0% on roasted coffee. MPF 0.3464%. None of it quoted yet."),
    ("3PL receiving, storage, pick-pack", "Per pallet and per order. No rate card exists - the 3PL is not selected."),
    ("Wire and FX cost on roaster payments", "Mercury wire fees plus the EUR/USD spread. The FX assumption itself is an open decision in the brief."),
]
for label, note in var:
    c1 = cs.cell(row=r, column=1, value=label); c1.font = bold; c1.alignment = top_wrap; c1.border = box
    c4 = cs.cell(row=r, column=4, value=note); c4.font = cell_font; c4.alignment = top_wrap; c4.border = box
    for c in (2, 3):
        cs.cell(row=r, column=c).border = box
    cs.row_dimensions[r].height = 32
    r += 1

for col, w in zip("ABCD", (52, 16, 16, 92)):
    cs.column_dimensions[col].width = w

ds = wb.create_sheet("Decisions")
ds["A1"] = "Decisions on add-on solutions - made, and pending"
ds["A1"].font = title_font
ds["A2"] = "The decision layer behind the inventory. LOCKED means a Standard or CLAUDE.md sec9 records it. OPEN means it will force a spec revision if it lands the wrong way."
ds["A2"].font = sub_font

DH = ["State", "Decision", "System(s)", "What was decided", "Why it matters", "Date", "Recorded in", "Owner", "Next action", "By when"]
for c, h in enumerate(DH, start=1):
    cell = ds.cell(row=4, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.border = box
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

D = [
["LOCKED","Subscription engine","Loop Subscriptions","Loop, replacing Recharge and Skio. Shopify-native selling plans plus Shopify Checkout, with a Loop-hosted portal rather than a hand-built pause/skip/swap engine.","Theme-level subscription code becomes native selling_plan_groups, so there is no rewrite risk from an engine swap.","2026-06-29","CLAUDE.md sec9; Store Operating Standards sec6"],
["LOCKED","Loop tier","Loop Subscriptions","Starter ($99/mo + 1.0%). Free is enough to TEST the design and not enough to RUN it; Pro at $399 is not needed.","Four named commitments sit above Free: dunning, cancellation flows, branded portal, subscription shipping rates. Pro was only ever wanted for its API, and contract rates turned out to be editable on Free.","2026-08-21","production_build_spec.md sec5.2.1, sec5.2.2"],
["LOCKED","Bundle app","Shopify Bundles (native)","Use Shopify's own Bundles app. Do NOT buy a third-party bundle app.","Paid apps earn their fee on mix-and-match and BOGO, none of which we need; and the two things no app does (component-derived facets, freshness gating) are ours to build either way.","2026-08-21","production_build_spec.md sec7.1"],
["LOCKED","Reviews backend","Judge.me","Judge.me free tier as a COLLECTION AND MODERATION BACKEND ONLY. Our own bespoke control renders the rating. Custom Questions ($15/mo) not needed.","It writes Shopify's standard review data, so the vendor is swappable - changing vendor changes the back office, not the storefront.","2026-08-20","production_build_spec.md sec6.1; Store Operating Standards sec13"],
["LOCKED","Shopify plan floor","Shopify","Basic is DISQUALIFIED - it includes zero extra staff accounts and the team is four people. Grow is the recommendation.","Nothing else in the design needs more: Functions run on all plans, account extensions run on all plans, and checkout extensibility was declined at about $24k/yr.","2026-08-21","production_build_spec.md sec10 checklist"],
["LOCKED","Checkout promo field","Shopify (plan)","Do NOT buy Plus to hide the checkout discount field. Amend the Standard instead - we issue NO discount codes at all, so the field is inert.","Verified on the dev store: field-level control below Plus covers only name, company, address line 2 and phone. The Checkout Editor lists the discount field but is advisory and non-interactive.","2026-07-25","Store Operating Standards v1.3 sec10"],
["LOCKED","Account surface","Customer-account UI extensions","The account page is NOT buildable in Liquid. A store created today runs new customer accounts only; /account redirects off-domain.","The POC had been modelling a surface we do not own. Business rules are untouched; the rendering surface and technique change.","2026-07-25","production_build_spec.md sec0, sec5.1"],
["LOCKED","Discount policy","Shopify Functions","No discount ever stacks. The applied rate is MAX() of everything the customer qualifies for. No codes are issued at all.","Tight per-bag margins - additive stacks could eat our lunch. Shopify already applies only the largest product discount per line off Plus, which hands us the rule for free.","2026-07-13","Store Operating Standards v1.2 sec3"],
["LOCKED","Email tool at launch","Shopify Email","Shopify Email free at launch; migrate to Klaviyo later.","Whichever platform is used must deliver campaign discounts as TIME-BOXED CUSTOMER TAGS, never codes or /discount/ links.","2026-06","00_PROJECT_BRIEF.md sec10, sec15"],
["LOCKED","Price maintenance","Spreadsheet + Shopify Flow + later a custom app","Phased. Spreadsheet-assisted at launch, Flow for Offerta aging, a lightweight approve/hold/defer app only when volume justifies it.","Landed cost x markup with approval governance is not a native Shopify feature, and building it now would delay the storefront for a back-office convenience.","2026-07-13","Store Operating Standards sec11, sec12.2"],
["LOCKED","Banking","Mercury","Mercury as the US business bank.","Chosen for Shopify integration and EUR/SWIFT wire capability for roaster payments in Italy.","2026-06","00_PROJECT_BRIEF.md sec2"],
["LOCKED","This sheet is the SINGLE HOME for open items","Governance","Open items had been living in THREE places: this sheet, CLAUDE.md sec10 checklist, and the follow-ups inside DECISIONS_LOG entries. From 2026-08-22 this sheet is canonical; the others point at it and must not restate it.","Same repair already applied twice at document level - brief sec12 on 2026-07-14 and brief sec10 today - now applied to the list of open items itself. The failure it prevents was visible the same day: the mailbox item sat OPEN in CLAUDE.md sec10 while already closed in reality. Filter State on this sheet for OPEN, GAP or CHANGING to get the live list.","2026-08-22","Steve + Code"],
["OPEN","Google Workspace tier and seat count","Google Workspace","Neither the plan tier nor the number of seats is recorded. Business Starter is assumed.","It is the only unknown left in the monthly cost line: $7/user/month on annual billing, so one seat is $7 and four is $28. Also decides whether Lucia, Asia and Lauren have company mailboxes at all, which bears on the alias-vs-Group call.","Open","Raised 2026-08-22"],
["OPEN","Packaging vs labeling ambiguity in the Roaster Guide","Roaster Guide","The packaging section requires bags compatible with U.S.-applied retail labeling; the labeling section says labels are applied at origin and Non rietichettiamo negli Stati Uniti. Deliberately deferred out of v8 rather than bundled.","Bundling an unsettled ambiguity into a painstakingly reviewed translation is how that translation gets silently broken. Needs deciding on its merits, then a v9 - not a quiet edit.","Open","Build spec sec15.1; deferred from v8"],
["OPEN","Rename _v6_pending_it - the name says draft, the file is approved","Roaster Guide","Crema_Italia_Roaster_Guide_v6_pending_it.html/.pdf still carries pending in its name although Lucia approved it on 2026-08-21.","Housekeeping, but it is the file every v6-to-v8 comparison is diffed against, so a misleading name costs someone a double-take at exactly the wrong moment.","Open","Build spec sec15.1"],
["SCHEDULED","Let the Namecheap SSL lapse","Namecheap","Decision is made - do not renew. The ACTION has a date: the certificate expires 12-NOV-2026 and must simply not be renewed.","Recorded as a scheduled action because the decision itself is RESOLVED and would otherwise drop off the live list, taking its dated action with it. Before the date, confirm nothing installed it by hand on a non-Namecheap server. Also check the order history for the cancelled duplicate 34236491.","Action due 12-NOV-2026","Namecheap account read 2026-08-22"],
["RESOLVED","Entitlement owner - DECIDED, Standard v1.15","Loop + Shopify Functions","DECIDED BY STEVE 2026-08-22, published as Store Operating Standards v1.15. The LOOP SELLING PLAN owns subscription lines - founder 12% and subscriber 10% as two separate plans - and a FUNCTION owns one-time lines, with appliesOnSubscription:false as the guard. Standard sec12.7 and sec12.8 close.","Closes the architecture break the spike found. Confirmed empirically first: a Function discount IS snapshotted onto the contract and DOES reach renewals (order #1002, contract #15302394080 - the contract carries the plan 12% and the Function 10% side by side). recurringCycleLimit is the control: 1 is first-order-only, N is N cycles, 0 is indefinite - so at 0 the compounding would be permanent rather than a first-order slip, which is why the guard matters. What survives unchanged: the rate is a snapshot taken at signup, so entitlement is CONTRACT STATE, not computed state.","2026-08-22","Standard v1.15; commits 6696cac, aec3107"],
["OPEN","Contract rates cannot be corrected in bulk - every remedy is per-contract","Loop","PROVEN 2026-08-22 during dev-store teardown: the probe discount object was DELETED and the subscription contract still showed that Function discount as active, usage limit intact. A contract snapshot is a COPY, not a reference - so amending or deleting a discount does nothing to anyone already subscribed.","THE STRONGEST FORM OF THE POINT STANDARD sec11 NOW RESTS ON, and it cuts both ways. GOOD: a founder 12% cannot be removed by a later change to the discount, deliberate or accidental, so Standard sec4 durable founder status gains a MECHANICAL guarantee rather than only a policy one. BAD: a wrong rate propagates permanently to everyone who signed up under it, and fixing the discount unwinds nothing. IT ALSO NARROWS AN EARLIER ESTIMATE: the Loop Starter-over-Pro recommendation was costed at 222 manual edits, capped because Founding Membership is capped - but that covered PROMOTIONS only. A rate error or policy change applied to existing subscribers is one manual edit per affected contract and is capped by nothing. It does not overturn Starter at expected volumes; the reasoning was just narrower than it looked. PRACTICAL RULE: treat a published selling-plan rate as effectively immutable for existing subscribers, and check it twice before the first signup rather than trusting that it can be fixed later.","Open","Commit 7eba722; dev-store teardown"],
["OPEN","Feeding Tampa Bay must confirm it can accept coffee","No-waste pledge / hero copy","Steve wrote to Feeding Tampa Bay 2026-08-22 to confirm they can accept our beans as a donation. The answer now gates live storefront copy.","THE DEPENDENCY IS NOT NEW, BUT POC21 MOVED IT INTO THE HERO. The Promise page has named Feeding Tampa Bay since POC3 - What doesn't sell goes to Feeding Tampa Bay is now the third sentence a visitor reads, so a no answer makes the most visible text on the site false rather than one line on a secondary page. WHY THEY MIGHT DECLINE, and it is not safety: coffee at 150 days is shelf-stable and low risk, but food banks commonly have policy on accepting goods past a printed best-by or quality date, and ours is donated precisely because it has passed ours. Ask about THAT specifically rather than about coffee in general. ALSO CONFIRM while asking: whether they will take whole bean in retail bags rather than bulk, what volumes and cadence suit them, and who receives it - the brief names Pete Lenhardt, plenhardt@feedingtampabay.org. AND THE TAX ANGLE: the brief claims the IRS 170(e)(3) enhanced deduction, which requires a qualified donee using the property for its exempt purpose, so their answer bears on that too. IF THEY DECLINE: the pledge itself survives - it is a real commitment and the differentiator - but the named recipient must change in the hero, the Promise page and the Offerta as-is copy together, and no other food bank has been approached.","Open","Steve, 2026-08-22; brief sec14"],
["OPEN","Win-back 15% cannot out-rank the standing rate on a subscription","Loop + Functions","Knowingly accepted gap in the v1.15 model. Win-back is the only campaign that could beat the standing subscriber rate on a subscription line, and under appliesOnSubscription:false it cannot. The top-up to MAX is DESIGNED AND UNBUILT.","Accepted deliberately rather than missed, which is the important distinction - but it is a live divergence from Standard sec3s MAX rule on exactly one path. Build the top-up, or accept that win-back does not apply to existing subscribers and say so where the rate is published.","Open","Standard v1.15; commit aec3107"],
["GAP","Duplicate SKUs give INDEPENDENT stock pools - overselling, not just mis-picking","Shopify / 3PL","C3 finding 2026-08-22: Shopify permits duplicate SKUs across products and gives each variant its OWN inventory item. Two products sharing one physical SKU therefore have independent stock. Adjusting one to 30 left the other at 40.","THIS ENLARGES sec13.9.2. The problem was framed as FIFO handing a full-price buyer the aged bag - a picking error. It is worse: the same physical stock is counted twice, so the risk is OVERSELLING. Candidate B (segregate by location only) does not address it. Feeds directly into the 3PL qualifying questions.","Open","Commit aec3107; build spec sec13.9.2"],
["RESOLVED","Function vs selling plan","Shopify Functions + Loop","ANSWERED 2026-08-22 by a real test order, not by reasoning. A discount Function DOES compound with the selling-plan price adjustment: 12% plan then a 10% Function billed $19.77 on a $24.95 variant, an effective 20.76%. combinesWith made no difference on any of the three discount classes.","The either/or worry is superseded and the outcome is better than expected: the Function is handed the PRE-PLAN base price and can see the subscription line, so it can top up to MAX instead of being excluded from subscription lines. That is what prevents MAX collapsing to the standing rate on every subscription. Standard sec12.7 answered yes in the same run - a Function reads customer tags and metafields.","2026-08-22","commit fa07ad3; dev-store test order"],
["OPEN","Benefit visibility","Shopify / theme","The subscriber benefit is INVISIBLE on a Shopify order - no discount line, just a lower price. Decide whether the theme renders it from base-vs-plan price.","Proven on order #1001: $24.95 billed at $21.96 with no discount line anywhere. Shopify's discount analytics will report zero discounts on subscription orders.","Open","CLAUDE.md sec10 item A3"],
["RESOLVED","Bundle inventory - native Bundles derive availability","Shopify Bundles","ANSWERED 2026-08-22 (B2). Native bundles derive availability from components, so the sec7.1 recommendation to use Shopify own Bundles rather than buy a third-party app stands.","The one ambiguity that could have overturned that recommendation is closed.","2026-08-22","Commit 8683f7d"],
["OPEN","Account branding","Customer-account UI extensions","Does the accounts and checkout branding editor offer Marcellus?","Two minutes. Decides whether the hosted account surface diverges from the storefront on TYPE as well as layout.","Open","CLAUDE.md sec10 item B1"],
["OPEN","Review records","Judge.me","Does metaobject syndication require Shop eligibility, and does a syndicated review populate author with the customer reference?","The second answer decides whether the palate-match join is free or has to be built. Blast radius is small by design - only the review DETAIL view needs individual records.","Open","CLAUDE.md sec10 item C4"],
["OPEN","Flow reliability","Shopify Flow","Shopify Flow for Offerta aging has never been tested. Liquid date comparison is unreliable in Flow.","The whole phased price-tool approach rests on Flow doing the aging transition.","Open","CLAUDE.md sec10 item C5"],
["OPEN","Customer service tool","Shopify Inbox / Gorgias / email only","Three-way choice, never made.","Coupled to the mailbox question - POC9's contact form routes to info@ (which EXISTS, verified 2026-08-22), support@ and contact@ (which do not).","Open","00_PROJECT_BRIEF.md sec18"],
["OPEN","Klaviyo migration trigger","Klaviyo","At what active-subscriber count do we migrate? Default 200.","Also carries a budget question for paid acquisition tooling that has never been answered.","Open","00_PROJECT_BRIEF.md sec18"],
["OPEN","Is Locksmith still needed?","Locksmith","Decided in the 2026-06 brief and never re-tested. The 48-hour Selezione early-access rule survives in the Standard, but the mechanism has not been re-examined.","Oldest app decision in the record. A customer tag plus native collection availability might do it without a paid app.","Open","00_PROJECT_BRIEF.md sec10; Store Operating Standards sec1"],
["OPEN","Affiliate app","Refersion / UpPromote / GoAffPro","Deferred post-launch, vendor unchosen.","The theme owns only the landing page and the application entry point.","Deferred","production_build_spec.md sec4"],
["OPEN","Referral reward and tooling","TBD","Both the reward FORM and the capture tooling are open. The old free-100g-bag reward is void.","100g exists only inside collections, so there is no standalone SKU to gift. And whatever is chosen must not require issuing a discount code.","Open","Store Operating Standards sec12.6"],
["OPEN","3PL","Tampa 3PL","Not selected. Three QUALIFYING questions must be asked before commercial terms.","Offerta-vs-fresh segregation of the same SKU, our packing slip vs their paperwork, and per-order printed inserts. A 3PL that fails the third cannot fulfil Sorpresa at all.","Open","Store Operating Standards sec12.9"],
["OPEN","Freight forwarder","Italian freight forwarder","Not selected.","Also blocks a piece of site content - Partner 1 on the About page is this vendor.","Open","00_PROJECT_BRIEF.md sec8"],
["OPEN","Customs broker","US customs broker","Not selected.","Files entry, ISF, FDA Prior Notice and manages any FDA hold.","Open","00_PROJECT_BRIEF.md sec8"],
["OPEN","Accounting system","QuickBooks / Wave / spreadsheets","Never chosen. Everything runs on OneDrive spreadsheets today.","Fine at pre-launch volume; not fine once orders and roaster payables both start moving.","Open","00_PROJECT_BRIEF.md sec18"],
["RESOLVED","Domain registrar and DNS","Namecheap","REGISTRAR AND DNS ARE BOTH NAMECHEAP - confirmed by Steve. Nameservers are Namecheap BasicDNS. Domain and PremiumDNS add-on both paid through 29-APR-2027.","Closes the whole gap. The zone carries the storefront records AND the MX records for the address the FDA holds as US Agent contact, so it is the one system whose failure takes revenue and regulatory correspondence down together.","Confirmed 2026-08-22","Steve, 2026-08-22. No prior document named any of it."],
["RESOLVED","Flip to PremiumDNS","Namecheap","DONE 2026-08-22. Nameservers switched to pdns1/pdns2.registrar-servers.com and VERIFIED by querying the new nameservers directly - every record intact, both nameservers agreeing, public resolvers already updated, site HTTP 200. No downtime.","Closes the paid-for-and-unused finding. Done pre-launch, which was the cheapest possible moment - no orders to lose. Method note: the verification queried DNS directly rather than reading the admin screen, on the same principle this project already applies to deployment state - live output beats a document, and it beats a settings page too.","2026-08-22","This session; verified by live DNS query"],
["RESOLVED","Email authentication - closed for now (Steve)","Google Workspace / Shopify / DNS","SPF, DKIM and DMARC are all in place and all passing on live mail. Steve's call 2026-08-22: rest on the subject. No further email-authentication work is queued.","Recording this as a DECISION rather than leaving the thread dangling, so a later reader does not read the remaining item as an oversight. THE ONE THING DELIBERATELY NOT DONE: DMARC is published at p=none, which monitors and enforces nothing. Moving to p=quarantine is the natural next step once a few weeks of reports look clean, and Shopify's admin carries a standing DMARC prompt that links to general documentation rather than to a status check - it is not reporting a gap. Not now, by choice.","2026-08-22","Steve"],
["RESOLVED","SPF record on cremaitalia.com","Google Workspace / DNS","ADDED AND VERIFIED 2026-08-22. v=spf1 include:_spf.google.com ~all, resolving at the authoritative nameservers and at public resolvers, as exactly one spf1 record.","Found and closed the same day. Google's own DMARC report had shown spf: none -> fail, with DKIM carrying every message alone. Lookup cost is 1 of 10 - _spf.google.com currently returns a flat ip4/ip6 record with no nested includes.","2026-08-22","Live DNS query; Google DMARC aggregate report"],
["RESOLVED","Shopify sender authentication","Shopify Email","BOTH sets verified healthy end to end 2026-08-22. Set A (SendGrid, p662) and Set B (Mailgun, p581), the latter with DKIM hosts nested under the mailerast subdomain. All four DKIM chains resolve to live RSA keys.","No fault. Two earlier conclusions here were wrong: a guessed host name returning NXDOMAIN was read as breakage, and a grep for the string DKIM1 matched a HOSTNAME rather than a key, reporting healthy records as dead. Both are the same failure class already logged in this project - a check that matches text rather than meaning can report the exact opposite of the truth. Verify against the key material (p=MII), not the label.","2026-08-22","DNS verification against authoritative nameservers"],
["RESOLVED","Sender-set cleanup - decided: LEAVE BOTH","Shopify Email","Set A (SendGrid, selector a7q) is the live signing identity, proven from a real message. The cleanup question - whether to delete Set B - is CLOSED with the decision to leave both in place. The Shopify admin route was tried and is a dead end: it hides the expected DNS record list once a domain is authenticated, so there is nothing to compare against the zone.","The records cost nothing, they are inert, and they consume no apex SPF lookups because Shopify authenticates via a mailer subdomain. The evidence for deleting Set B would be one signed message, not a statement of configuration - and deleting the wrong set breaks DKIM on transactional mail, which Standard sec8.1 makes the customer's only price record. Revisit only if Shopify itself ever shows one identity as retired.","2026-08-22","Message headers + Shopify admin read"],
["CORRECTED","DMARC external reporting","DNS / asymplat.biz","EARLIER CLAIM WAS WRONG. The RFC 7489 authorisation record genuinely does not exist (NXDOMAIN, verified under the correct name with a control query) - but Steve DOES receive the reports. Google sends them regardless.","The record is missing, the conclusion drawn from it was not. Adding cremaitalia.com._report._dmarc.asymplat.biz is HARDENING so other reporters that enforce the check also send - not the repair of a live failure. Method lesson: I first queried the wrong record name, then drew a confident conclusion from an absence. An absence is the weakest evidence there is, and it deserves a control query before anything is built on it.","Corrected 2026-08-22","Steve's DMARC report; re-verified DNS query"],
["GAP","WHO files FDA Prior Notice? The record contradicts itself","FDA / customs broker","Brief section 4 says Prior Notice is filed BY CREMA ITALIA before every shipment from roaster-supplied data. Sections 8 and 9 say the CUSTOMS BROKER handles entry, ISF, FDA Prior Notice and hold management. Both are present tense, in the same document.","Found 2026-08-22 while mapping the inbound data flows. Must be settled before the first shipment: a Prior Notice that each party assumes the other filed is a shipment held at the border, and the failure appears at the worst possible moment. Also a broker-selection question - it changes what we are buying from them.","Open","Brief sec4 vs sec8/sec9; Data Flows A6"],
["CORRECTED","Packing list lot columns - smaller than first stated","Roaster / 3PL","An earlier entry called this the highest-value fix in the inbound chain. WRONG. Roaster Guide v7 already requires lot number and roast date on the BAG LABEL, on every MASTER CARTON label, and in the recall protocol. The narrow truth: the packing list is specified only as one line per SKU per bag size, with no lot or roast-date columns.","Lot data is not missing - it arrives physically on cartons rather than in transcribable form, so someone at the 3PL reads and keys it. Adding two columns is still worth doing, because it turns a manual transcription into a document handoff, but it is a v8 tidy rather than a risk. Steve was right that v7 captured this; the build spec 15.1 note is accurate but reads worse than the reality when quoted alone.","Corrected 2026-08-22","Read of Roaster Guide v7 _us and _it"],
["RESOLVED","Roast date format - CLOSED by Roaster Guide v8","Roaster / 3PL / FDA","Verified across BOTH v7 editions: no date-format wording, no ISO 8601 reference, no worked date example. The guide requires a roast date on the bag and the carton and never says how to write it.","THIS IS THE SHARPER RISK, and it is the one the packing-list item was masking. 03/07/2026 is 3 July to an Italian roaster and 7 March to a US warehouse. That single ambiguity breaks FIFO picking (Standard 5.4), the 90-day freshness gate and the Offerta transition - silently, and in the direction of shipping coffee that is four months older than the system believes. Build spec 15.2 already places this in the PRO-FORMA LABEL step rather than the pre-boarding guide, which is the right home; the decision is DD-MMM-YYYY per 2026-08-21. It just has not been written into anything a roaster reads.","Closed 2026-08-22","Roaster Guide v8 _it and _us, both editions verified"],
["GAP","Nobody owns the FDA biennial re-registration","FDA / roasters","Every partner roaster must renew its Foreign Food Facility registration BIENNIALLY. No owner, calendar or tracker is named anywhere.","A lapsed registration blocks import. It will fail SILENTLY two years after a roaster signs, when nobody is thinking about it and the relationship feels settled - which is exactly the kind of deadline that needs an owner rather than an intention. Lucia owns the roaster relationship and is the natural candidate.","Open","Brief sec4; Data Flows A5"],
["RESOLVED","Namecheap SSL - attached to nothing; let it lapse","Namecheap / Shopify","Read from the account 2026-08-22. PositiveSSL 34236492 is issued for cremaitalia.com and expires 12-NOV-2026, managed by the shared-hosting tool - but there is NO hosting on the account, and Shopify refuses third-party certificates on every plan. It protects nothing. A second cert, 34236491, is CANCELED and was never issued. DECISION: do not renew.","Shopify provisions and auto-renews a Let's Encrypt certificate at its load balancer, free on every plan, so there is no configuration in which a purchased certificate protects the storefront. DATE CORRECTION: the 29-APR-2027 renewal covers the domain and PremiumDNS only - the certificate runs on its own clock. Check the order history for a possible double purchase, and confirm nothing installed it by hand on a non-Namecheap server before letting it go.","2026-08-22","Namecheap account read; live TLS check"],
["LOCKED","External maintenance systems anticipated; shared hosting DECLINED","Architecture","Steve asked whether staging systems will be stood up outside Shopify - roaster onboarding, product onboarding - and whether the Namecheap hosting should have been kept for an admin.cremaitalia.com subdomain. YES to external systems; NO to shared hosting.","Four such systems are already specified: Product Onboarding and Roaster Onboarding (build spec sec15, sec13.4.3 - roaster metaobjects are API-writable, which is what makes an application form possible), the price-maintenance approve/hold/defer queue (Standard sec11/sec12.2), and the reorder-rate job (sec6.1). BUT THEY SPLIT THREE WAYS AND ONLY ONE WANTS A HOST. Roaster records are already specified as METAOBJECTS and a staged product is a DRAFT PRODUCT - Shopify supplies storage, permissions and admin UI free, so rebuilding either outside would create a second catalogue. The scheduled job wants a scheduler, not a web server. Only an approval queue with its own state wants hosting, and its natural form is a Shopify EMBEDDED ADMIN APP, inheriting staff authentication rather than needing a second user system. Namecheap shared hosting is cPanel/LAMP and a poor fit for a Node service needing env secrets, webhooks and deploy-from-git; a managed Node host fits. THE KEY SEPARATION: admin.cremaitalia.com is a DNS RECORD, NOT HOSTING. The zone is now controlled at Namecheap PremiumDNS and can point anywhere in ninety seconds, so keeping the hosting was never what supplied the subdomain. THE TRAP TO GUARD: a staging system must be strictly UPSTREAM and write-once - it assembles a proposal, pushes to Shopify, and Shopify is canonical from that instant. If it keeps editing SKUs that already exist in Shopify, that is the two-homes drift Review A removed on 2026-08-20, reintroduced somewhere nobody diffs. TIMING: build nothing yet. No roaster has signed and the entire catalogue is fixture data, so a pipeline built now would be designed against imagined inputs. Standard sec12.2 already locked spreadsheet-first with a volume trigger. Onboard the first real roaster MANUALLY, write down what actually happens, then automate what hurt - probably SKU assignment and label generation. The pro-forma label may need no service at all: the repo already has a WeasyPrint pipeline with brand fonts and render gates.","2026-08-22","Steve + Code; build spec sec15, Standard sec11/sec12.2"],
["SUPERSEDED","Flip to PremiumDNS, or stop paying for it?","Namecheap","PremiumDNS is SUBSCRIBED but NOT ACTIVE - it only answers queries once the nameservers point at pdns1/pdns2.registrar-servers.com, and they currently point at BasicDNS. Two honest options: switch and use it, or drop it.","Paying for a service that is not serving. PremiumDNS buys a 100% uptime SLA (BasicDNS has no guarantee), DDoS protection and 30+ anycast nodes. DNSSEC is NOT the differentiator - Namecheap supports it on both. RECOMMENDATION: flip, and do it PRE-LAUNCH, when there are no orders to lose. CAUTION: do not assume records auto-copy; make both zones identical BEFORE switching, because during propagation resolvers may hit either nameserver set.","Open - raised by Steve 2026-08-22","Namecheap KB; this session"],
["RESOLVED","Email provider and mailboxes","Google Workspace","PROVIDER IS GOOGLE WORKSPACE, and the mailbox picture is now read off the admin rather than inferred. Primary user steve.roberts@cremaitalia.com (created 22-JUN-2026) carries five aliases: info@, sroberts@, steve@, usagent@, roasters@. Missing: support@ and contact@. Tier and seat count still to confirm.","CORRECTS CLAUDE.md sec10, which states that info@ does not exist - it does, and that is a present-tense claim in a live document. usagent@ confirmed live matters independently: it is the address the FDA holds as US Agent contact.","Confirmed 2026-08-22","Workspace admin; supersedes CLAUDE.md sec10"],
["OPEN","support@ and contact@ - alias or Group?","Google Workspace","Both are free to add (30 aliases per user, every plan). The open call is HOW. Every address today is an alias into one inbox.","A Google Group is also free, needs no licence, and can include EXTERNAL addresses - so support@ as a Group could reach Lauren without buying her a seat. Standard sec9 puts customer service with Crema Italia, and the team is four people. Also worth asking whether three addresses landing in one inbox earn their keep, or whether the POC form should offer two choices rather than three.","Open","Raised 2026-08-22"],
["OPEN","Is the FDA number a Dialpad number?","Dialpad","The FDA holds +1-813-376-4821 as the US Agent contact for every roaster registration. Confirm whether that is the Dialpad line or a personal mobile.","If it is Dialpad, the subscription lapsing breaks a REGULATORY contact of record, and the remedy runs through the FDA rather than a billing page. Different risk class from losing calls.","Open","00_PROJECT_BRIEF.md sec4; raised 2026-08-22"],
["OPEN","Does the site publish a phone number?","Dialpad / theme","The storefront publishes NO phone number - no tel: link anywhere in the theme. Dialpad exists to publish numbers; decide whether the site should carry one.","Either the site should show a number, or the seat count should reflect what is actually used. Also feeds JSON-LD contactPoint, currently unemitted.","Open","Theme grep 2026-08-22"],
["OPEN","Two cloud document stores","OneDrive vs Google Workspace","Google Workspace includes Drive, and the business also pays for OneDrive/M365.","Not automatically waste - the Cowork lane and every Standard render path is built on local OneDrive paths. But worth a deliberate look rather than discovering it on a renewal notice.","Open","Raised 2026-08-22"],
["GAP","Analytics stack","GA4 / Shopify Analytics","No decision recorded anywhere.","And Shopify's own discount reporting is structurally blind to the subscriber benefit, so it cannot be the answer on its own.","Open","No governing document"],
["GAP","Reorder-job runtime","In-house scheduled job","Reorder rate and palate-match need a scheduled job writing product metafields. No runtime has been chosen.","A quiet infrastructure decision hiding inside a content feature.","Open","production_build_spec.md sec6.1"],
]

# Owner / Next action / By when - only live items carry them; settled rows are dashed out.
TRACK = {
 "Entitlement owner": ("Steve", "Decide whether the rate lives on the Loop contract or in a Function, then amend Standard sec11/sec12.8.", "Before production build"),
 "Benefit visibility": ("Steve + Code", "Decide whether the theme renders the subscriber benefit from base-vs-plan price.", "Before launch"),
 "Bundle inventory": ("Code", "Dev store: build a two-component bundle, place a test order, watch component stock.", "10 minutes, any time"),
 "Account branding": ("Steve", "Settings > Checkout > Configurations > Edit - does the branding editor offer Marcellus?", "2 minutes, any time"),
 "Review records": ("Steve", "Ask Judge.me support the two questions in production_build_spec sec6.1.", "Before reviews go live"),
 "Flow reliability": ("Code", "Test a Flow scheduled trigger with Run-code date arithmetic on the dev store.", "Before production build"),
 "Customer service tool": ("Steve", "Choose: Shopify Inbox, Gorgias, or email only.", "Before launch"),
 "Klaviyo migration trigger": ("Steve", "Confirm the active-subscriber count that triggers migration (default 200).", "Post-launch"),
 "Is Locksmith still needed?": ("Code", "Test whether a customer tag plus native collection availability delivers the 48h Selezione gate without the app.", "Before launch"),
 "Affiliate app": ("Steve", "Choose a vendor when the affiliate programme is actually scheduled.", "Post-launch"),
 "Referral reward and tooling": ("Steve", "Decide the reward form - it must not require issuing a discount code.", "Before any referral discount"),
 "3PL": ("Steve", "Shortlist and ask the three QUALIFYING questions in Standard sec12.9 before commercial terms.", "LAUNCH BLOCKING"),
 "Freight forwarder": ("Lucia + Steve", "Select from the Italian_Freight_Forwarder_Selection study.", "LAUNCH BLOCKING"),
 "Customs broker": ("Steve", "Select, and settle the Prior Notice ownership question in the same conversation.", "LAUNCH BLOCKING"),
 "Accounting system": ("Steve", "Choose QuickBooks, Wave, or stay on spreadsheets deliberately.", "Before first orders"),
 "WHO files FDA Prior Notice? The record contradicts itself": ("Steve", "Settle with the broker at selection; correct brief sec4 or sec8/sec9 so they agree.", "LAUNCH BLOCKING"),
 "Nobody owns the FDA biennial re-registration": ("Lucia", "Assign an owner and put each roaster renewal date in a calendar at signing.", "At first roaster signing"),
 "support@ and contact@ - alias or Group?": ("Steve", "Decide alias vs Google Group. Deferred as trivial 2026-08-22.", "Before customer service goes live"),
 "Is the FDA number a Dialpad number?": ("Steve", "Confirm whether +1-813-376-4821 is the Dialpad line or a personal mobile.", "Before launch"),
 "Does the site publish a phone number?": ("Steve + Code", "Decide whether the storefront carries a tel: link; feeds JSON-LD contactPoint.", "Before launch"),
 "Two cloud document stores": ("Steve", "Deliberate look at OneDrive vs Google Drive overlap - not a fix, a decision.", "At next renewal"),
 "Analytics stack": ("Steve", "Choose an analytics stack. Note Shopify discount reporting is blind to subscriber benefits.", "Before launch"),
 "Reorder-job runtime": ("Code", "Evaluate a scheduled GitHub Action - zero new vendor, repo already exists.", "When reviews reach the floor"),
 "Google Workspace tier and seat count": ("Steve", "Read the tier and seat count off the Workspace billing page.", "5 minutes, any time"),
 "Packaging vs labeling ambiguity in the Roaster Guide": ("Steve + Lucia", "Decide which statement governs, then issue v9. Do not quiet-edit v8.", "Before first roaster signs"),
 "Rename _v6_pending_it - the name says draft, the file is approved": ("Cowork", "Rename both .html and .pdf off the pending marker.", "Housekeeping"),
 "Contract rates cannot be corrected in bulk - every remedy is per-contract": ("Steve + Code", "Before the first real signup, verify each published selling-plan rate. Decide whether a bulk-correction path is ever needed (that would be Loop Pro API territory).", "Before first signup"),
 "Feeding Tampa Bay must confirm it can accept coffee": ("Steve", "Await their reply; ask specifically about goods past a printed quality date, retail bags vs bulk, and volume/cadence.", "Before launch - hero copy depends on it"),
 "Win-back 15% cannot out-rank the standing rate on a subscription": ("Steve + Code", "Build the top-up to MAX, or publish that win-back does not apply to existing subscribers.", "Before win-back emails ship"),
 "Duplicate SKUs give INDEPENDENT stock pools - overselling, not just mis-picking": ("Steve + Code", "Re-open the three sec13.9.2 candidates against overselling, not just mis-picking; take it into the 3PL questions.", "LAUNCH BLOCKING"),
 "Let the Namecheap SSL lapse": ("Steve", "Do NOT renew. Confirm first that nothing installed it on a non-Namecheap server.", "12-NOV-2026"),
}
D = [d + list(TRACK.get(d[1], ("-", "-", "-"))) for d in D]

for d in D:
    ds.append(d)

for row in ds.iter_rows(min_row=5, max_row=4 + len(D), min_col=1, max_col=len(DH)):
    for cell in row:
        cell.font = cell_font
        cell.alignment = top_wrap
        cell.border = box
    row[1].font = bold
    st = str(row[0].value)
    fill = {"LOCKED": "FFEFF6EF", "PART RESOLVED": "FFEFF6EF", "RESOLVED": "FFEFF6EF",
            "CHANGING": "FFFDF6EA", "SUPERSEDED": "FFF6F3EC", "CORRECTED": "FFFDF6EA", "OPEN": "FFF6F3EC", "GAP": "FFFBEDEC"}.get(st)
    if fill:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
    ds.row_dimensions[row[0].row].height = 62

for col, w in zip("ABCDEFGHIJ", (12, 26, 30, 62, 62, 14, 34, 16, 56, 22)):
    ds.column_dimensions[col].width = w
ds.freeze_panes = "C5"
ds.auto_filter.ref = "A4:J%d" % (4 + len(D))

lg = wb.create_sheet("Legend & Sources")
lg["A1"] = "How to read this workbook, and where every claim came from"
lg["A1"].font = title_font
state = {"r": 3}

def block(title, items):
    r = state["r"]
    lg.cell(row=r, column=1, value=title).font = Font(name=ARIAL, size=11, bold=True, color=ESP)
    r += 1
    for a, b in items:
        ca = lg.cell(row=r, column=1, value=a); ca.font = bold; ca.alignment = top_wrap
        cb = lg.cell(row=r, column=2, value=b); cb.font = cell_font; cb.alignment = top_wrap
        lg.row_dimensions[r].height = 30
        r += 1
    state["r"] = r + 1

block("Status values", [
 ("LIVE", "Running and paid for (or free) today."),
 ("INSTALLED", "Installed on the free Partners development store for validation, not yet on the production store."),
 ("SELECTED", "Decided and recorded in a Standard or the brief, not yet installed or paid for."),
 ("BUILD", "We write it ourselves. No vendor and no subscription - developer time."),
 ("PHASED", "Deliberately staged: a manual method now, an app later, with a named trigger."),
 ("CANDIDATE", "Named as a likely answer, but no decision has been recorded."),
 ("PENDING DECISION", "An explicit open choice. Someone has to decide."),
 ("NOT SELECTED", "A vendor category with no vendor. Several of these are launch-blocking."),
 ("GAP", "Something the business already depends on that is recorded nowhere."),
 ("PART RESOLVED", "Decisions sheet only. Part of a gap has been closed and the remainder is named."),
])

block("Cost conventions", [
 ("Blank cost cell", "Genuinely unknown, or not applicable. Never zero-by-assumption - a blank here is a question, not a saving."),
 ("$0", "Verified free at our usage level."),
 ("Monthly Cost Now", "What is actually being paid today."),
 ("Monthly Cost at Launch", "What the decided stack costs once the store opens. Excludes every PENDING and NOT SELECTED line."),
 ("Other Costs", "Percentages, per-unit and per-shipment costs. On this business these are larger than the subscriptions."),
])

block("Primary sources used", [
 ("CLAUDE.md", "The theme repo's project memory: the sec9 decision log (67 dated entries), sec10 current state, and the open-item checklists."),
 ("docs/production_build_spec.md", "sec5.1 account extensions, sec5.2 / 5.2.1 / 5.2.2 the Loop findings, sec6.1 reviews, sec7.1 bundle apps, sec10 the plan choice."),
 ("docs/standards/store-operating-standards.md", "v1.14. sec3 discounts, sec4 Founding Member, sec6 Loop subscriptions, sec8 fulfilment, sec11 tooling, sec12 open decisions, sec13 reviews."),
 ("00_PROJECT_BRIEF.md (OneDrive)", "sec2 entity and banking, sec4 FDA, sec8 logistics, sec9 US operations, sec10 commerce stack and app list, sec13 shipping, sec15 email flows, sec17 launch checklist, sec18 decisions pending."),
 ("OneDrive Operations tree", "Fulfillment Selection studies (3PL, forwarder, broker), Accounting, the In Italy FDA pack, and the Coordination decisions log."),
 ("Claude Code session transcripts", "48 MB of session history scanned for vendor names, to catch systems discussed but never written into a document."),
 ("Vendor pricing, checked 2026-08-22", "Judge.me (judge.me/pricing), Locksmith (locksmith.guide/policies/pricing), Klaviyo, Anthropic Claude Max. The Loop and Shopify figures come from our own verified records, not from marketing pages."),
])

block("Two things worth knowing before quoting any number", [
 ("Live output beats a document", "This repo's standing rule, learned the hard way three separate times. Every figure here is only as good as its last check - re-verify before committing money."),
 ("Loop is the larger platform cost", "It was locked as the engine on 2026-06-29 and nobody priced it until 2026-08-21, eight weeks later. A cost line should be a required field on any platform decision."),
])

lg.column_dimensions["A"].width = 34
lg.column_dimensions["B"].width = 118


# ---------------------------------------------------------------- Sheet: Data Flows
fl = wb.create_sheet("Data Flows")
fl["A1"] = "Data exchange flows - who sends what to whom, by what mechanism, how often"
fl["A1"].font = title_font
fl["A2"] = ("Two clusters, and the asymmetry between them is the headline. INBOUND (supply) is low-frequency, high-value, "
            "document-based and human-verified - a pallet every 6 to 10 weeks. OUTBOUND (demand) is high-frequency, "
            "low-value-per-event and API-driven - every order. Automating inbound would be waste; outbound API capability "
            "is a selection criterion. Compiled 2026-08-22 from the build spec, the Standards and the project brief.")
fl["A2"].font = sub_font
fl["A3"] = ("THE INTERCHANGE KEY IS THE SKU. Build spec 13.9.1: it travels - the roaster, the freight forwarder, the 3PL, a "
            "packing slip, a scanner, a bag label all see it; none of them can see a metafield. Every EXTERNAL party integrates "
            "on the SKU string; every INTERNAL relationship runs on metaobject references. That is already decided and it "
            "governs every row below.")
fl["A3"].font = Font(name=ARIAL, size=9, italic=True, color=ESP)

FH = ["Flow", "Cluster", "From", "To", "What moves", "Mechanism today",
      "Mechanism target", "Cadence", "Initiator", "Status", "Notes and risk"]
for c, h in enumerate(FH, start=1):
    cell = fl.cell(row=5, column=c, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.border = box
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

F = [
["A1","INBOUND","Roaster","Crema Italia","SKU proposal: name and tagline, origin, varietal, process, roast level, 3-6 tasting notes, back-story, brewing methods, 2-3 hi-res photographs (the Roaster Guide SKU back-story payload).","Undefined. Email and attachments in practice.","Tokenised web form writing a coffee metaobject as DRAFT, photographs included, made ACTIVE on acceptance (build spec 13.4.3).","Per SKU. Event-driven, low volume.","Roaster","PROCESS NOT DEFINED","Build spec 15.2 says this process does not exist today and its inputs are scattered across the Roaster Guide. Step 2, review and accept, has NO written criteria beyond excluding flavored coffee and permitting declared decaf."],
["A2","INBOUND","Roaster","Crema Italia","Roaster application: identity, town, region, contact, website, logo and photographs.","Intro letter and guide, then email.","Same tokenised-form mechanism as A1 - writes a roaster metaobject as Draft. Described once in 13.4.3 rather than twice.","Once per roaster.","Lucia Calo (Operations Manager - Italy)","PROCESS NOT DEFINED","Ends when the roaster exists as a roaster metaobject. The guide, intro letter and vendor sheet already exist in OneDrive Operations/In Italy. Italian is the document of record."],
["A3","INBOUND","Crema Italia","Roaster","The ASSIGNED SKU CODE (TRRRPPPPSS) plus the label template.","Document and email.","Generated from the metaobject references, never typed (13.9.1), then issued with the label template.","Per SKU, after acceptance.","Crema Italia","SPECIFIED, NOT BUILT","Mandatory since Roaster Guide v7: the SKU must appear on the label the roaster applies at origin, so the warehouse can pick individual bags. This is the reverse flow people forget - the roaster cannot label correctly until we send this."],
["A4","INBOUND","Crema Italia <-> Roaster","(both)","Pro-forma label, generated and audited before the first production run.","Not built.","Generated from agreed data. The repo ALREADY has a WeasyPrint pipeline with brand fonts and render gates - the same shape as producing a label.","Per SKU, once, before first production.","Crema Italia","SPECIFIED, NOT BUILT","This is where the ROAST DATE FORMAT belongs and it must be unambiguous: 03/07/2026 is 3 July to an Italian roaster and 7 March to a US warehouse, which would break FIFO picking and freshness gating. Use DD-MMM-YYYY, per the 2026-08-21 decision."],
["A5","INBOUND","Roaster","FDA","Foreign Food Facility registration, with biennial renewal.","FDA Industry Systems (FURLS) web form. Roaster submits; Crema Italia walks them through.","No change - this is a government system and cannot be automated.","Once per roaster, then BIENNIAL renewal.","Roaster, guided by Crema Italia","LIVE (process exists)","GAP: nobody is named as the owner of tracking the BIENNIAL renewal per roaster. A lapsed registration blocks import, and it will fail silently two years after a roaster signs, when nobody is thinking about it. Needs a calendar owner."],
["A6","INBOUND","Crema Italia OR the customs broker","FDA","Prior Notice, before every shipment, built from roaster-supplied data.","Manual submission.","Undecided - see the contradiction opposite.","Per shipment.","DISPUTED","CONTRADICTION IN THE RECORD","THE RECORD DISAGREES WITH ITSELF. Brief section 4 says Prior Notice is filed BY CREMA ITALIA before every shipment. Brief sections 8 and 9 say the CUSTOMS BROKER handles entry, ISF, FDA Prior Notice and hold management. Both are present-tense. This must be settled before the first shipment - a Prior Notice that both parties assume the other filed is a shipment held at the border."],
["A7","INBOUND","Roaster","Freight forwarder (AFF)","Pickup booking, pallet ready notice, dimensions and weights. FCA Incoterm, roastery-to-airport haul in our cost.","Email or phone.","No change expected. Volume does not justify integration.","Per shipment, every 6-10 weeks.","Roaster","VENDOR NOT SELECTED","Air freight only. FCO or MXP to TPA, forwarder choice per shipment."],
["A8","INBOUND","Roaster","Crema Italia, then forwarder and broker","Documents travelling with the pallet: English commercial invoice (our template), English PACKING LIST one line per SKU per bag size, ICO Certificate of Origin, ASL health certificate if applicable.","Documents, our templates.","Same - but see the risk. No phytosanitary certificate is required for roasted coffee.","Per shipment.","Roaster","MOSTLY COVERED - see correction","CORRECTED 2026-08-22 AFTER READING v7. An earlier version of this row called the missing packing-list columns the highest-value fix in the inbound chain. That was OVERSTATED. Roaster Guide v7 requires lot and roast data in THREE places: the BAG LABEL mandatory elements carry Lot/batch code and roast date plus the Crema Italia SKU code; each MASTER CARTON is labelled with SKU, bag size, bag count, gross/net weight, LOT NUMBER, ROAST DATE and country of origin; and the recall protocol requires lot number and roast date within 24 hours. Cartons are one SKU and one bag size each, so lot data arrives complete and well-organised. WHAT IS ACTUALLY MISSING is narrower: the packing list is specified only as one line per SKU per bag size, so it carries no lot or roast-date COLUMNS. The consequence is not missing data - it is that lot data arrives PHYSICALLY, on cartons, rather than in transcribable form. Someone at the 3PL reads cartons and keys it. Fine at eight pallets a year; still worth two columns, because it converts a manual transcription into a document handoff and removes a dependency on 3PL diligence."],
["A9","INBOUND","Freight forwarder","Customs broker","Pre-alert, air waybill, the document set from A8.","Forwarder to broker, email or portal.","No change.","Per shipment.","Forwarder","VENDOR NOT SELECTED","Carriers seen: American FCO-TPA direct, Lufthansa via FRA, IAG via LHR, Delta via ATL/JFK, SWISS via ZRH."],
["A10","INBOUND","Customs broker","CBP and FDA","Entry filing, ISF where applicable, Prior Notice (see A6), FDA hold management.","Broker EDI/ABI - their system, not ours.","No change. We never touch this directly.","Per shipment.","Broker","VENDOR NOT SELECTED","HS 0901.21 roasted not decaffeinated, 0901.22 decaffeinated. DUTY 0%. MPF 0.3464% applies; HMF is ocean-only so not applicable to air."],
["A11","INBOUND","Customs broker","Crema Italia","Entry summary, duties and fees, brokerage charges - the components that complete LANDED COST.","Invoice and PDF.","No change expected at this volume.","Per shipment.","Broker","VENDOR NOT SELECTED","Feeds landed_cost_usd, which is LOCKED AT RECEIPT and drives every retail price through cost x Markup[shelf,size]. An error here propagates to every price on the shelf."],
["A12","INBOUND","Crema Italia","Mercury -> Roaster","Payment for goods, EUR by SWIFT wire.","Manual initiation in the Mercury web app.","Mercury publishes an API (read plus payment initiation). NEVER EVALUATED - no integration designed.","Per purchase order / shipment.","Crema Italia","LIVE, MANUAL","Mercury was chosen partly FOR EUR/SWIFT capability. The FX assumption itself is still an open decision in the brief - single locked rate, spot, or forward-hedged, and who carries the risk. eur_usd_rate is locked per lot for historical cost tracking."],
["A13","INBOUND","Mercury","Accounting","Statements and transaction history.","Manual export. Everything reconciles in OneDrive spreadsheets today.","Undecided - the accounting system itself is unchosen (QuickBooks vs Wave vs spreadsheets).","Monthly, or ad hoc.","Crema Italia","NO INTEGRATION","Fine at pre-launch volume. Stops being fine once orders and roaster payables are both moving. Nothing connects Mercury to Shopify or to any ledger."],
["A14","INBOUND","Crema Italia","Shopify","Cost per variant (unitCost) and the computed retail price.","Spreadsheet computes, human types into Shopify. Standard 12.2 locks this as the launch method.","Phase 3 only: a lightweight approve/hold/defer app, WHEN VOLUME JUSTIFIES IT.","Per receipt, every 6-10 weeks.","Crema Italia","LOCKED AS MANUAL, BY DESIGN","unitCost must be set via REST, and there is a staff-permission catch (build spec onboarding order, step 5). Price is COMPUTED from cost x markup and never typed. Deliberately manual at launch - this is a decision, not a gap."],
["A15","INBOUND -> OUTBOUND","3PL","Shopify","Goods receipt: quantities per SKU, and the LOT record (roast date, lot code) created per receipt.","Unknown - depends entirely on the unselected 3PL.","Lot metaobject created per receipt and referenced from the product; product before lot, one direction only.","Per pallet, every 6-10 weeks.","3PL","VENDOR NOT SELECTED","THE SEAM WHERE INBOUND BECOMES OUTBOUND, and the one inbound flow that genuinely wants a system rather than a document. Depends on A8 carrying roast date and lot code."],
["B1","OUTBOUND","Customer","Shopify","Order: lines, quantities, shipping address, gift flag (cart attribute), selling plan if subscribed.","Native Shopify. Realtime.","No change - this is Shopify core.","Per order.","Customer","NATIVE","The gift flag rides to the order as a CART ATTRIBUTE because checkout takes no custom fields below Plus (Standard 8.2)."],
["B2","OUTBOUND","Loop <-> Shopify","(both)","Subscription contracts, pre-scheduled recurring orders, and lifecycle events.","Loop is Shopify-native: selling plans plus Shopify Checkout. Realtime via webhooks.","Same.","Per signup, then per cycle (4/6/8 weeks). Contracts pre-schedule future orders.","Loop","LIVE ON DEV STORE, VERIFIED","VERIFIED 2026-08-21 on order #1001: the contract stores base price, discount percent and plan as its OWN fields and had already pre-scheduled five future orders. THE RATE IS CONTRACT STATE, NOT A RULE EVALUATED PER ORDER."],
["B3","OUTBOUND","Loop","Shopify customer tags","Subscription-state events driving entitlement: actively-shipping, in-60-day-grace, lapsed.","Loop webhooks plus Shopify Flow. Realtime on event.","Same. NEVER maintained by the theme - a customer can cancel from an email link and never touch storefront UI.","Event-driven.","Loop","SPECIFIED, NOT BUILT","Standard 11. This is the authoritative server-side downgrade path. Note the architecture break: discount Functions do NOT re-run on recurring orders, so entitlement for orders 2..n is contract state, which makes it Loop's job."],
["B4","OUTBOUND","Shopify Flow","Shopify","Scheduled maintenance: Offerta transition flag, unpublish at the end of the freshness window, founding-member and active-roccia tagging.","Not built.","Daily scheduled trigger, using the RUN CODE action for date arithmetic - date comparison in Flow Liquid conditions is unreliable and fails silently.","Daily.","Shopify Flow","SPECIFIED, NEVER TESTED","Two jobs and only the second should be fully automatic: FLAG a SKU approaching the Offerta threshold for a human to decide how much stock to split; UNPUBLISH at the end of the window. The second is the hard stop the no-waste pledge depends on."],
["B5","OUTBOUND","Shopify","3PL","Fulfilment request: order, lines, SKUs, quantities, ship-to, service level.","Unknown - depends on the unselected 3PL.","App, API or EDI. THIS IS THE ONE FLOW WHERE REALTIME API CAPABILITY IS A SELECTION CRITERION, because it fires per order.","Per order. Realtime.","Shopify","VENDOR NOT SELECTED","Contrast with every inbound flow: this happens per order rather than per pallet, so batch or email-and-CSV becomes a bottleneck immediately rather than eventually."],
["B6","OUTBOUND","Shopify","3PL","Packing slip (NO monetary fields, Standard 8.1), the component BOM for Sorpresa collections, and the gift-card insert flag.","Not built.","Shopify packing slip template, which we control. BOM from the bundle. Gift flag from the order attribute.","Per order.","Shopify","SPECIFIED","THREE QUALIFYING QUESTIONS for the 3PL, all in Standard 12.9: (0) how do you segregate an Offerta lot from fresh stock of the SAME SKU in the same bin, given FIFO would hand the aged bag to a full-price buyer; (a) do you print OUR packing slip or insert your own paperwork - their own paperwork breaks 8.1 and nothing on our side prevents it; (b) can you insert a printed card varied per order. A 3PL that fails (b) cannot fulfil Sorpresa at all."],
["B7","OUTBOUND","3PL","Shopify","Fulfilment confirmation, tracking number, inventory adjustments, and the ORDER-LINE LOT STAMP at fulfilment.","Unknown.","API writeback. The lot stamp is what makes recall traceability work.","Per order. Realtime.","3PL","VENDOR NOT SELECTED","Build spec onboarding order, step 11. Without the lot stamp there is no way to answer which customers received a given lot."],
["B8","OUTBOUND","Shopify","Customer","Transactional email: order confirmation, shipping, account activation, plus the eight marketing flows.","LIVE. Sent as info@cremaitalia.com, DKIM-signed with selector a7q via SendGrid.","Same, plus Shopify Email or Klaviyo for the marketing flows.","Per order and per campaign.","Shopify","LIVE AND VERIFIED","Standard 8.1 makes the emailed confirmation the customer's ONLY price record, because nothing inside the package shows a price. That raises the stakes on deliverability from marketing inconvenience to the customer never receiving their receipt."],
["B9","OUTBOUND","Shopify Payments","Mercury","Payouts - settled card revenue.","Native Shopify, scheduled batch. One-way.","No change.","Per payout schedule.","Shopify","LIVE","LOOP NEVER TOUCHES THE MONEY, and this is worth stating because people assume it does. Loop bills THROUGH Shopify Checkout, so funds flow Shopify Payments to Mercury like any other order. Loop's 1% is an app charge on the Shopify invoice, not a deduction from the payout."],
["B10","OUTBOUND","Shopify","Accounting","Revenue, fees, refunds, payout reconciliation.","No integration. Nothing exists.","Undecided - the accounting system is unchosen.","Monthly.","Crema Italia","NO INTEGRATION","Pairs with A13. Both ends of the money story are currently manual, and the accounting system that would join them has never been chosen."],
]

for row in F:
    fl.append(row)

fw = [8, 16, 20, 20, 40, 30, 34, 20, 18, 22, 70]
for i, w in enumerate(fw, start=1):
    fl.column_dimensions[get_column_letter(i)].width = w

flast = 5 + len(F)
for row in fl.iter_rows(min_row=6, max_row=flast, min_col=1, max_col=len(FH)):
    for cell in row:
        cell.font = cell_font
        cell.alignment = top_wrap
        cell.border = box
    row[0].font = bold
    row[0].alignment = ctr
    cluster = str(row[1].value)
    tint = "FFF3F7FA" if cluster.startswith("INBOUND") else "FFFAF6F0"
    if "->" in cluster:
        tint = "FFF6F1F7"
    st = str(row[9].value)
    if "NOT SELECTED" in st or "CONTRADICTION" in st or "NOT DEFINED" in st or "NO INTEGRATION" in st:
        tint = "FFFBEDEC"
    elif st.startswith("LIVE") or st.startswith("NATIVE"):
        tint = "FFEFF6EF"
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=tint)
    fl.row_dimensions[row[0].row].height = 110

fl.freeze_panes = "C6"
fl.auto_filter.ref = "A5:%s%d" % (get_column_letter(len(FH)), flast)


wb.save(OUT)
print("saved:", OUT, "| inventory rows:", len(R), "| decision rows:", len(D))
